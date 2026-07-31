#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="168A50")
SUBHEADER_FILL = PatternFill("solid", fgColor="E6F4EA")
FAIL_FILL = PatternFill("solid", fgColor="FFE6E6")
BLOCK_FILL = PatternFill("solid", fgColor="FFF4D6")
PASS_FILL = PatternFill("solid", fgColor="E6F4EA")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN_BORDER = Border(
    left=Side(style="thin", color="B7C5BD"),
    right=Side(style="thin", color="B7C5BD"),
    top=Side(style="thin", color="B7C5BD"),
    bottom=Side(style="thin", color="B7C5BD"),
)

CASE_SHEET_CANDIDATES = [
    "精确自动化执行用例",
    "最终自动化可执行用例",
    "自动化模拟方案",
    "用户视角测试场景",
]

CASE_COLUMNS = [
    "用例ID",
    "优先级",
    "产品模块",
    "子功能",
    "测试场景",
    "前置条件",
    "测试数据",
    "执行入口/Selector",
    "执行步骤",
    "预期结果",
    "成功判定",
    "失败判定",
    "证据要求",
    "自动化Runner",
    "执行层级",
    "每轮必跑",
    "来源ID",
    "来源类型",
    "备注",
]

RESULT_COLUMNS = [
    "本轮执行状态",
    "本轮结果分类",
    "本轮实际结果",
    "本轮问题描述",
    "本轮证据目录",
    "本轮关键截图",
    "本轮报告路径",
    "本轮LLM复核状态",
    "本轮LLM复核文件",
    "本轮执行时间",
]

STATUS_CN = {
    "passed": "通过",
    "failed": "失败",
    "blocked": "阻塞",
    "skipped": "跳过",
    "needs_llm_review": "需LLM复核",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="QBot casebook export/result writer")
    sub = parser.add_subparsers(dest="command", required=True)

    export = sub.add_parser("export-cases")
    export.add_argument("--casebook", required=True)
    export.add_argument("--output", required=True)
    export.add_argument("--profile", default="mandatory", choices=["mandatory", "full", "all"])
    export.add_argument("--sheet", default="")
    export.add_argument("--case", default="")
    export.add_argument("--offset", type=int, default=0)
    export.add_argument("--limit", type=int, default=0)

    write = sub.add_parser("write-results")
    write.add_argument("--casebook", required=True)
    write.add_argument("--summary", required=True)
    write.add_argument("--output", required=True)
    return parser.parse_args()


def clean(value) -> str:
    if value is None:
        return ""
    return ILLEGAL_CHARACTERS_RE.sub("", str(value)).strip()


def excel_safe(value) -> str:
    if value is None:
        return ""
    return ILLEGAL_CHARACTERS_RE.sub("", str(value))


def find_case_sheet(wb):
    sheets = find_case_sheets(wb)
    if sheets:
        return sheets[0]
    raise ValueError("未找到包含“用例ID”的测试用例 sheet")


def find_case_sheets(wb):
    candidate_sheets = []
    for name in CASE_SHEET_CANDIDATES:
        if name in wb.sheetnames:
            ws = wb[name]
            header_row = find_header_row(ws)
            if header_row and has_case_id_header(header_map(ws, header_row)):
                candidate_sheets.append(ws)
    if candidate_sheets:
        return candidate_sheets
    sheets = []
    for ws in wb.worksheets:
        header_row = find_header_row(ws)
        if header_row and has_case_id_header(header_map(ws, header_row)):
            sheets.append(ws)
    return sheets


def has_case_id_header(headers: dict[str, int]) -> bool:
    return "用例ID" in headers or "原用例ID" in headers


def find_header_row(ws) -> int:
    for row in range(1, min(ws.max_row or 1, 10) + 1):
        values = [clean(ws.cell(row, col).value) for col in range(1, (ws.max_column or 1) + 1)]
        has_case_id = "用例ID" in values or "原用例ID" in values
        if has_case_id and ("测试场景" in values or "测试目标" in values):
            return row
    return 0


def header_map(ws, header_row: int) -> dict[str, int]:
    return {
        clean(ws.cell(header_row, col).value): col
        for col in range(1, (ws.max_column or 1) + 1)
        if clean(ws.cell(header_row, col).value)
    }


def row_value(ws, row: int, headers: dict[str, int], name: str) -> str:
    return clean(ws.cell(row, headers[name]).value) if name in headers else ""


def row_value_any(ws, row: int, headers: dict[str, int], names: list[str]) -> str:
    for name in names:
        value = row_value(ws, row, headers, name)
        if value:
            return value
    return ""


def parse_json_cell(value: str, *, field: str, case_id: str):
    raw = clean(value)
    if not raw:
        return None
    try:
        return json.loads(raw)
    except json.JSONDecodeError as error:
        raise ValueError(
            f"{case_id} 的 {field} 不是合法 JSON："
            f"line={error.lineno}, column={error.colno}, {error.msg}"
        ) from error


def split_case_ids(value: str) -> set[str]:
    return {item.strip() for item in re.split(r"[,，\s]+", value or "") if item.strip()}


def is_selected(row: dict, profile: str, wanted: set[str]) -> bool:
    if wanted:
        return row["id"] in wanted
    if profile in {"full", "all"}:
        return True
    return row.get("mandatory") == "是"


def infer_case_kind(row: dict) -> str:
    runner = row.get("runner", "")
    runner_text = runner.lower().strip()
    runner_tail = runner_text.split("/")[-1].strip() if "/" in runner_text else runner_text
    if "ui+conversation" in runner_tail or "ui + conversation" in runner_tail:
        return "ui+conversation"
    if runner_tail == "attachment" or "attachment" in runner_tail:
        return "attachment"
    if runner_tail == "auth" or "auth" in runner_tail:
        return "auth"
    if runner_tail == "conversation" or "conversation" in runner_tail:
        return "conversation"
    if runner_tail == "ui" or re.search(r"(^|\s)ui($|\s)", runner_tail):
        return "ui"

    primary_text = "\n".join([
        row.get("module", ""),
        row.get("submodule", ""),
        row.get("scenario", ""),
        row.get("test_data", ""),
    ])
    scenario_text = "\n".join([
        primary_text,
        row.get("precondition", ""),
        row.get("steps", ""),
        row.get("selectors", ""),
    ])
    if re.search(r"询问.*token|环境变量|系统提示词|拒绝泄露|敏感信息|内部信息保护", scenario_text, re.I):
        return "conversation"
    if (
        re.search(r"登录与账号|OAuth|登录成功|登录失败|退出登录|重新登录|鉴权|授权页|取消授权|refresh\s*token|会话恢复", scenario_text, re.I)
        or row.get("module", "") == "登录与账号"
    ):
        return "auth"
    if (
        re.search(r"成果区|成果预览|任务成果|产物|artifact|聊天正文不混入|安全沙箱", primary_text, re.I)
        and not re.search(r"附件|上传|文件上传|读取.*文件", primary_text, re.I)
    ):
        return "conversation"
    if re.search(r"附件|上传|文件上传|读取.*文件|图片附件|图片上传|PDF附件|Word附件|Excel附件|PPT附件|多模态附件", primary_text, re.I):
        return "attachment"
    if re.search(r"Agent|回复|收到回复|会话|输入消息|多轮|总结|摘要|提问|问候|上下文|对话|长文本输入", primary_text, re.I):
        return "conversation"
    if re.search(r"设置|菜单|入口|按钮|切换|筛选|弹窗|页签|收起|展开|搜索", scenario_text, re.I):
        return "ui"
    return "ui"


def export_cases(args: argparse.Namespace) -> None:
    casebook = Path(args.casebook)
    wb = load_workbook(casebook, data_only=True)
    case_sheets = find_case_sheets(wb)
    if not case_sheets:
        raise ValueError("未找到包含“用例ID”的测试用例 sheet")
    if args.sheet:
        requested = clean(args.sheet)
        matched = [ws for ws in case_sheets if clean(ws.title) == requested]
        if not matched:
            available = ", ".join(ws.title for ws in case_sheets)
            raise ValueError(f"未找到指定测试用例 sheet：{requested}；可用 sheet：{available}")
        case_sheets = matched
    wanted = split_case_ids(args.case)

    rows = []
    export_sheets = export_case_sheets(wb, case_sheets)
    sheet_names = [ws.title for ws in case_sheets]
    for ws in export_sheets:
        header_row = find_header_row(ws)
        headers = header_map(ws, header_row)
        for row_idx in range(header_row + 1, (ws.max_row or header_row) + 1):
            case_id = row_value_any(ws, row_idx, headers, ["用例ID", "原用例ID"])
            if not case_id:
                continue
            row = {
                "id": case_id,
                "priority": row_value(ws, row_idx, headers, "优先级") or "P1",
                "module": row_value(ws, row_idx, headers, "产品模块") or ws.title,
                "submodule": row_value(ws, row_idx, headers, "子功能"),
                "scenario": row_value(ws, row_idx, headers, "测试场景") or row_value(ws, row_idx, headers, "测试目标"),
                "precondition": row_value(ws, row_idx, headers, "前置条件"),
                "test_data": row_value(ws, row_idx, headers, "测试数据"),
                "selectors": row_value(ws, row_idx, headers, "执行入口/Selector") or row_value(ws, row_idx, headers, "入口/Selector"),
                "steps": row_value(ws, row_idx, headers, "执行步骤") or row_value(ws, row_idx, headers, "自动化执行步骤"),
                "expected_result": row_value(ws, row_idx, headers, "预期结果") or row_value(ws, row_idx, headers, "断言标准"),
                "success_criteria": row_value(ws, row_idx, headers, "成功判定"),
                "failure_criteria": row_value(ws, row_idx, headers, "失败判定") or row_value(ws, row_idx, headers, "失败/阻塞判定"),
                "evidence_required": row_value(ws, row_idx, headers, "证据要求"),
                "runner": row_value(ws, row_idx, headers, "自动化Runner") or row_value(ws, row_idx, headers, "推荐命令/Runner"),
                "execution_level": row_value(ws, row_idx, headers, "执行层级") or row_value(ws, row_idx, headers, "执行方式"),
                "mandatory": row_value(ws, row_idx, headers, "每轮必跑"),
                "source_id": row_value(ws, row_idx, headers, "来源ID"),
                "source_type": row_value(ws, row_idx, headers, "来源类型"),
                "note": row_value(ws, row_idx, headers, "备注") or row_value(ws, row_idx, headers, "维护备注"),
                "user_journey": row_value(ws, row_idx, headers, "用户旅程"),
                "blocking_level": row_value(ws, row_idx, headers, "阻断等级"),
                "pipeline_policy": row_value(ws, row_idx, headers, "流水线策略"),
                "second_review_required": row_value(ws, row_idx, headers, "二次复核要求"),
                "case_type": row_value(ws, row_idx, headers, "用例类型"),
                "core_domain": row_value(ws, row_idx, headers, "核心域"),
                "batch_size": row_value(ws, row_idx, headers, "批次大小"),
                "initialization_policy": row_value(ws, row_idx, headers, "初始化策略"),
                "contract_version": row_value(ws, row_idx, headers, "契约版本"),
                "automation_protocol": row_value(ws, row_idx, headers, "自动化协议"),
                "evidence_schema_version": row_value(ws, row_idx, headers, "证据Schema版本"),
                "risk_domain": row_value_any(ws, row_idx, headers, ["风险域", "风险分类"]),
                "oracle_type": row_value_any(ws, row_idx, headers, ["判定Oracle", "Oracle类型", "判定类型"]),
                "deterministic": row_value_any(ws, row_idx, headers, ["确定性", "是否确定性"]),
                "repeat_policy": row_value_any(ws, row_idx, headers, ["重复策略", "稳定性重复策略"]),
                "required_fixture": row_value_any(ws, row_idx, headers, ["必需Fixture", "环境与数据Fixture"]),
                "hard_gate": row_value_any(ws, row_idx, headers, ["硬门禁", "一票否决"]),
                "cleanup_policy": row_value_any(ws, row_idx, headers, ["清理策略", "数据清理策略"]),
                "version_scope": row_value_any(ws, row_idx, headers, ["版本范围", "适用版本"]),
                "known_bug_link": row_value_any(ws, row_idx, headers, ["历史Bug", "已知Bug链接"]),
                "production_signal": row_value_any(ws, row_idx, headers, ["生产观测指标", "上线观测指标"]),
                "contract_version": row_value_any(ws, row_idx, headers, ["契约版本", "Case契约版本"]),
                "product_baseline": row_value_any(ws, row_idx, headers, ["latest-main基线", "产品基线"]),
                "migration_disposition": row_value_any(ws, row_idx, headers, ["迁移处置", "重构处置"]),
                "visible_action_contract": row_value_any(ws, row_idx, headers, ["用户可见动作契约", "可见动作契约"]),
                "state_readback_contract": row_value_any(ws, row_idx, headers, ["状态读回契约", "独立读回契约"]),
                "required_evidence_roles": row_value_any(ws, row_idx, headers, ["证据角色", "必需证据角色"]),
                "forbidden_shortcuts": row_value_any(ws, row_idx, headers, ["禁止捷径", "禁止的捷径"]),
                "selector_contract": row_value_any(ws, row_idx, headers, ["Selector契约", "选择器契约"]),
                "identity_contract": row_value_any(ws, row_idx, headers, ["身份完整性", "发布身份契约"]),
                "trusted_review_contract": row_value_any(ws, row_idx, headers, ["可信复核契约", "可信评审规则"]),
                "case_type": row_value_any(ws, row_idx, headers, ["用例类型", "Case类型"]),
                "automation_protocol": row_value_any(ws, row_idx, headers, ["自动化协议", "执行协议"]),
                "initialization_profile": row_value_any(ws, row_idx, headers, ["初始化策略", "初始化Profile"]),
                "batch_size": row_value_any(ws, row_idx, headers, ["批次大小", "流水线批次大小"]),
                "action_plan_json": row_value_any(ws, row_idx, headers, ["动作计划JSON", "Action Plan JSON"]),
                "turns_json": row_value_any(ws, row_idx, headers, ["会话轮次JSON", "Turns JSON"]),
                "capability_policy_json": row_value_any(ws, row_idx, headers, ["能力抽样策略JSON", "Capability Policy JSON"]),
                "assertion_contract_json": row_value_any(ws, row_idx, headers, ["精准断言JSON", "Assertion Contract JSON"]),
                "evidence_schema_version": row_value_any(ws, row_idx, headers, ["证据Schema版本", "证据协议版本"]),
                "state_fixture_contract": row_value_any(ws, row_idx, headers, ["状态Fixture契约", "Fixture契约"]),
                "api_event_oracle": row_value_any(ws, row_idx, headers, ["API/事件Oracle", "API事件Oracle"]),
                "design_baseline_id": row_value_any(ws, row_idx, headers, ["设计基线ID"]),
                "design_baseline_file": row_value_any(ws, row_idx, headers, ["设计基线文件"]),
                "design_baseline_sha256": row_value_any(ws, row_idx, headers, ["设计基线SHA-256"]),
                "visual_comparison_contract": row_value_any(ws, row_idx, headers, ["视觉比对契约"]),
                "viewport_contract": row_value_any(ws, row_idx, headers, ["视口契约"]),
                "accessibility_contract": row_value_any(ws, row_idx, headers, ["无障碍契约"]),
                "branch_coverage": row_value_any(ws, row_idx, headers, ["分支覆盖"]),
                "traceability_tags": row_value_any(ws, row_idx, headers, ["可追溯标签"]),
                "sheet": ws.title,
                "row_number": row_idx,
            }
            row["action_plan"] = parse_json_cell(
                row_value(ws, row_idx, headers, "动作计划JSON"),
                field="动作计划JSON",
                case_id=case_id,
            )
            row["conversation_turns"] = parse_json_cell(
                row_value(ws, row_idx, headers, "会话轮次JSON"),
                field="会话轮次JSON",
                case_id=case_id,
            )
            row["capability_sampling"] = parse_json_cell(
                row_value(ws, row_idx, headers, "能力抽样策略JSON"),
                field="能力抽样策略JSON",
                case_id=case_id,
            )
            row["precise_assertions"] = parse_json_cell(
                row_value(ws, row_idx, headers, "精准断言JSON"),
                field="精准断言JSON",
                case_id=case_id,
            )
            row["evidence_roles"] = [
                item.strip()
                for item in re.split(
                    r"[,，;；|、\n]+",
                    row_value(ws, row_idx, headers, "证据角色")
                    or row.get("evidence_required", ""),
                )
                if item.strip()
            ]
            row["production_metadata_explicit"] = all(row.get(field) for field in [
                "risk_domain",
                "oracle_type",
                "deterministic",
                "repeat_policy",
                "required_fixture",
                "hard_gate",
                "cleanup_policy",
                "version_scope",
                "production_signal",
            ])
            explicit_kind = {
                "conversation": "conversation",
                "attachment": "attachment",
                "artifact": "conversation",
                "skill_lifecycle": "ui+conversation",
                "skill_use": "ui+conversation",
                "expert_lifecycle": "ui+conversation",
                "expert_use": "ui+conversation",
                "mcp_lifecycle": "ui+conversation",
                "mcp_use": "ui+conversation",
                "recovery": "ui+conversation",
                "auth_recovery": "auth",
                "run_initialization": "ui",
                "task_lifecycle": "ui+conversation",
                "project_lifecycle": "ui+conversation",
                "project_automation": "ui+conversation",
                "knowledge_lifecycle": "ui+conversation",
                "memory_lifecycle": "ui+conversation",
                "settings_lifecycle": "ui",
                "host_integration": "ui+conversation",
                "security_privacy": "ui+conversation",
                "performance_capacity": "ui+conversation",
            }.get(row["case_type"])
            row["kind"] = explicit_kind or infer_case_kind(row)
            if is_selected(row, args.profile, wanted):
                rows.append(row)

    if args.offset and args.offset > 0:
        rows = rows[args.offset :]

    if args.limit and args.limit > 0:
        rows = rows[: args.limit]

    output = Path(args.output)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(
        json.dumps(
            {
                "casebook": str(casebook),
                "sheets": sheet_names,
                "profile": args.profile,
                "selected_count": len(rows),
                "cases": rows,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )


def export_case_sheets(wb, case_sheets):
    # Smoke casebooks intentionally contain a total sheet plus per-module duplicate sheets.
    # Execute only the total sheet; write-results still updates all case sheets.
    if "冒烟测试总表" in wb.sheetnames:
        ws = wb["冒烟测试总表"]
        header_row = find_header_row(ws)
        if header_row and ("用例ID" in header_map(ws, header_row) or "原用例ID" in header_map(ws, header_row)):
            return [ws]
    return case_sheets


def normalize_status(status: str) -> str:
    return STATUS_CN.get(clean(status), clean(status))


def style_cell(cell, fill=None, font=None):
    cell.alignment = WRAP
    cell.border = THIN_BORDER
    if fill:
        cell.fill = fill
    if font:
        cell.font = font


def ensure_result_columns(ws, header_row: int, headers: dict[str, int]) -> dict[str, int]:
    next_col = (ws.max_column or 0) + 1
    for name in RESULT_COLUMNS:
        if name not in headers:
            ws.cell(header_row, next_col).value = name
            headers[name] = next_col
            next_col += 1
    for col in range(1, (ws.max_column or 1) + 1):
        style_cell(ws.cell(header_row, col), HEADER_FILL, WHITE_FONT)
    return headers


def key_screenshot(result: dict) -> str:
    shots = result.get("screenshots_flat") or []
    if not shots:
        screenshots = result.get("screenshots") or {}
        if isinstance(screenshots, dict):
            for value in screenshots.values():
                if isinstance(value, str):
                    shots.append(value)
    return shots[-1] if shots else ""


def result_indexes(results: list[dict]) -> tuple[dict[str, dict], dict[str, list[dict]]]:
    by_exact = {}
    by_id = {}
    for item in results:
        case_id = clean(item.get("id"))
        if not case_id:
            continue
        sheet = clean(item.get("sheet"))
        row_number = clean(item.get("row_number"))
        if sheet and row_number:
            by_exact[f"{sheet}::{row_number}::{case_id}"] = item
        by_id.setdefault(case_id, []).append(item)
    return by_exact, by_id


def write_results(args: argparse.Namespace) -> None:
    casebook = Path(args.casebook)
    summary = json.loads(Path(args.summary).read_text(encoding="utf-8"))
    output = Path(args.output)
    output.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(casebook, output)

    wb = load_workbook(output)
    case_sheets = find_case_sheets(wb)
    if not case_sheets:
        raise ValueError("未找到包含“用例ID”的测试用例 sheet")
    by_exact, by_id = result_indexes(summary.get("results", []))
    run_time = summary.get("ended_at") or datetime.now().isoformat()

    result_sheets = {
        clean(item.get("sheet"))
        for item in summary.get("results", [])
        if clean(item.get("sheet"))
    }
    for ws in case_sheets:
        if result_sheets and clean(ws.title) not in result_sheets:
            continue
        header_row = find_header_row(ws)
        headers = ensure_result_columns(ws, header_row, header_map(ws, header_row))
        for row_idx in range(header_row + 1, (ws.max_row or header_row) + 1):
            case_id_col = headers.get("用例ID") or headers.get("原用例ID") or 1
            case_id = clean(ws.cell(row_idx, case_id_col).value)
            if not case_id:
                continue
            result = by_exact.get(f"{ws.title}::{row_idx}::{case_id}")
            if result is None and len(by_id.get(case_id, [])) == 1:
                result = by_id[case_id][0]
            values = {}
            if result:
                values = {
                    "本轮执行状态": normalize_status(result.get("status")),
                    "本轮结果分类": result.get("result_category") or "",
                    "本轮实际结果": result.get("actual_result") or result.get("reason") or result.get("conclusion") or "",
                    "本轮问题描述": result.get("problem_description") or "",
                    "本轮证据目录": result.get("case_dir") or "",
                    "本轮关键截图": key_screenshot(result),
                    "本轮报告路径": result.get("case_report") or "",
                    "本轮LLM复核状态": result.get("llm_review", {}).get("status", ""),
                    "本轮LLM复核文件": result.get("llm_review", {}).get("prompt_file", ""),
                    "本轮执行时间": run_time,
                }
            else:
                values = {
                    "本轮执行状态": "未执行",
                    "本轮结果分类": "",
                    "本轮实际结果": "",
                    "本轮问题描述": "",
                    "本轮证据目录": "",
                    "本轮关键截图": "",
                    "本轮报告路径": "",
                    "本轮LLM复核状态": "",
                    "本轮LLM复核文件": "",
                    "本轮执行时间": "",
                }
            for name, value in values.items():
                cell = ws.cell(row_idx, headers[name])
                cell.value = excel_safe(value)
                fill = None
                if name == "本轮执行状态":
                    if value == "通过":
                        fill = PASS_FILL
                    elif value == "失败":
                        fill = FAIL_FILL
                    elif value == "阻塞":
                        fill = BLOCK_FILL
                style_cell(cell, fill)
        apply_sheet_style(ws, header_row)
    write_summary_sheets(wb, summary)
    wb.save(output)


def apply_sheet_style(ws, header_row: int = 1) -> None:
    for row in ws.iter_rows():
        for cell in row:
            style_cell(cell)
            if cell.row == header_row:
                style_cell(cell, HEADER_FILL, WHITE_FONT)
    ws.freeze_panes = ws.cell(header_row + 1, 1)
    for col in range(1, (ws.max_column or 1) + 1):
        letter = get_column_letter(col)
        width = 12
        for row in range(1, min(ws.max_row or 1, 80) + 1):
            value = ws.cell(row, col).value
            if value is not None:
                width = max(width, min(len(str(value)) + 2, 52))
        ws.column_dimensions[letter].width = width


def remove_sheet(wb, name: str) -> None:
    if name in wb.sheetnames:
        del wb[name]


def append_row(ws, values, header=False):
    ws.append([excel_safe(value) for value in values])
    row = ws.max_row
    for col in range(1, len(values) + 1):
        style_cell(ws.cell(row, col), HEADER_FILL if header else None, WHITE_FONT if header else None)


def write_summary_sheets(wb, summary: dict) -> None:
    for name in ["自动化执行汇总", "自动化执行明细", "Bug候选清单", "阻塞清单"]:
        remove_sheet(wb, name)

    counts = summary.get("counts", {})
    ws = wb.create_sheet("自动化执行汇总", 0)
    append_row(ws, ["指标", "值"], header=True)
    rows = [
        ["执行状态", summary.get("status", "")],
        ["执行目录", summary.get("run_dir", "")],
        ["用例源", summary.get("casebook", "")],
        ["结果Excel", summary.get("result_excel", "")],
        ["开始时间", summary.get("started_at", "")],
        ["结束时间", summary.get("ended_at", "")],
        ["总用例数", counts.get("total", 0)],
        ["通过", counts.get("passed", 0)],
        ["失败", counts.get("failed", 0)],
        ["阻塞", counts.get("blocked", 0)],
        ["需要LLM复核", counts.get("needs_llm_review", 0)],
    ]
    for row in rows:
        append_row(ws, row)
    apply_sheet_style(ws, 1)

    detail = wb.create_sheet("自动化执行明细")
    detail_headers = ["用例ID", "模块", "测试场景", "优先级", "状态", "结果分类", "实际结果", "证据目录", "关键截图", "报告路径", "LLM复核"]
    append_row(detail, detail_headers, header=True)
    for result in summary.get("results", []):
        append_row(detail, [
            result.get("id", ""),
            result.get("module", ""),
            result.get("scenario", "") or result.get("title", ""),
            result.get("priority", ""),
            normalize_status(result.get("status")),
            result.get("result_category", ""),
            result.get("actual_result", "") or result.get("reason", ""),
            result.get("case_dir", ""),
            key_screenshot(result),
            result.get("case_report", ""),
            result.get("llm_review", {}).get("status", ""),
        ])
    apply_sheet_style(detail, 1)

    bugs = wb.create_sheet("Bug候选清单")
    append_row(bugs, ["用例ID", "模块", "测试场景", "缺陷描述", "关键截图", "报告路径"], header=True)
    for result in summary.get("results", []):
        if result.get("status") == "failed" and result.get("result_category") == "bug":
            append_row(bugs, [
                result.get("id", ""),
                result.get("module", ""),
                result.get("scenario", "") or result.get("title", ""),
                result.get("problem_description", ""),
                key_screenshot(result),
                result.get("case_report", ""),
            ])
    apply_sheet_style(bugs, 1)

    blocked = wb.create_sheet("阻塞清单")
    append_row(blocked, ["用例ID", "模块", "测试场景", "阻塞原因", "证据目录", "关键截图", "报告路径"], header=True)
    for result in summary.get("results", []):
        if result.get("status") == "blocked":
            append_row(blocked, [
                result.get("id", ""),
                result.get("module", ""),
                result.get("scenario", "") or result.get("title", ""),
                result.get("actual_result", "") or result.get("reason", ""),
                result.get("case_dir", ""),
                key_screenshot(result),
                result.get("case_report", ""),
            ])
    apply_sheet_style(blocked, 1)


def main() -> None:
    args = parse_args()
    if args.command == "export-cases":
        export_cases(args)
    elif args.command == "write-results":
        write_results(args)


if __name__ == "__main__":
    main()
