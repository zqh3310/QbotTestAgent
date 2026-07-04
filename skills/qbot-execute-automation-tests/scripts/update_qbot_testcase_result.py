#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import shutil
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SCRIPT_DIR = Path(__file__).resolve().parent
SKILL_DIR = SCRIPT_DIR.parent
CASE_MAP_FILE = SKILL_DIR / "references" / "core-case-map.json"

HEADER_FILL = PatternFill("solid", fgColor="168A50")
SUBHEADER_FILL = PatternFill("solid", fgColor="DDEFE4")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="B7C5BD"),
    right=Side(style="thin", color="B7C5BD"),
    top=Side(style="thin", color="B7C5BD"),
    bottom=Side(style="thin", color="B7C5BD"),
)
WRAP = Alignment(wrap_text=True, vertical="top")


STATUS_CN = {
    "passed": "通过",
    "pass": "通过",
    "failed": "失败",
    "blocked": "阻塞",
    "skipped": "跳过",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Write QBot automation results back to Qbot_TestCase workbook.")
    parser.add_argument("--root", default=str(Path.cwd()), help="QbotTestAgent project root")
    parser.add_argument("--casebook", default=None, help="Source Qbot_TestCase.xlsx path")
    parser.add_argument("--run-dir", default=None, help="autoTest timestamp run directory")
    parser.add_argument("--summary", default=None, help="automation-run-summary.json path")
    parser.add_argument("--output", default=None, help="Output workbook path")
    parser.add_argument("--profile", default="mandatory", help="Core profile to mark")
    parser.add_argument("--mark-core-only", action="store_true", help="Only refresh core mandatory markers, without run results")
    return parser.parse_args()


def load_json(path: Path) -> dict:
    with path.open("r", encoding="utf-8") as file:
        return json.load(file)


def split_ids(value) -> list[str]:
    if value is None:
        return []
    return [item.strip() for item in str(value).replace("、", "\n").replace(",", "\n").splitlines() if item.strip()]


def ensure_sheet_removed(wb, name: str) -> None:
    if name in wb.sheetnames:
        del wb[name]


def normalize_status(status: str) -> str:
    return STATUS_CN.get(str(status or "").lower(), status or "")


def status_rank(status: str) -> int:
    status = str(status or "").lower()
    if status == "failed":
        return 3
    if status == "blocked":
        return 2
    if status in {"passed", "pass"}:
        return 1
    return 0


def aggregate_status(results: list[dict]) -> str:
    if not results:
        return ""
    worst = max(results, key=lambda item: status_rank(item.get("status")))
    return normalize_status(worst.get("status"))


def find_header(ws, candidates: list[str], default_row: int) -> int:
    for row in range(1, min(ws.max_row or 1, 8) + 1):
        values = [str(ws.cell(row, col).value or "").strip() for col in range(1, (ws.max_column or 1) + 1)]
        if any(candidate in values for candidate in candidates):
            return row
    return default_row


def header_map(ws, header_row: int) -> dict[str, int]:
    return {
        str(ws.cell(header_row, col).value or "").strip(): col
        for col in range(1, (ws.max_column or 1) + 1)
        if str(ws.cell(header_row, col).value or "").strip()
    }


def ensure_columns(ws, header_row: int, names: list[str]) -> dict[str, int]:
    headers = header_map(ws, header_row)
    next_col = (ws.max_column or 0) + 1
    for name in names:
        if name not in headers:
            ws.cell(header_row, next_col).value = name
            ws.cell(header_row, next_col).fill = HEADER_FILL
            ws.cell(header_row, next_col).font = WHITE_FONT
            ws.cell(header_row, next_col).alignment = WRAP
            ws.cell(header_row, next_col).border = THIN_BORDER
            headers[name] = next_col
            next_col += 1
    return headers


def style_sheet(ws, header_row: int = 1) -> None:
    for row in ws.iter_rows():
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = WRAP
            if cell.row == header_row:
                cell.fill = HEADER_FILL
                cell.font = WHITE_FONT
    ws.freeze_panes = ws.cell(header_row + 1, 1)
    for col in range(1, (ws.max_column or 1) + 1):
        letter = get_column_letter(col)
        max_len = 10
        for row in range(1, min(ws.max_row or 1, 80) + 1):
            value = ws.cell(row, col).value
            if value is not None:
                max_len = max(max_len, min(len(str(value)), 42))
        ws.column_dimensions[letter].width = max(12, min(max_len + 2, 46))


def style_existing_used_area(ws, header_row: int) -> None:
    for row in range(1, (ws.max_row or 1) + 1):
        for col in range(1, (ws.max_column or 1) + 1):
            cell = ws.cell(row, col)
            cell.border = THIN_BORDER
            cell.alignment = WRAP
            if row == header_row:
                cell.fill = HEADER_FILL
                cell.font = WHITE_FONT


def load_execution_mapping(wb) -> dict[str, dict]:
    mapping = {}
    if "自动化执行映射" not in wb.sheetnames:
        return mapping
    ws = wb["自动化执行映射"]
    headers = header_map(ws, 1)
    for row in range(2, (ws.max_row or 1) + 1):
        actual_id = ws.cell(row, headers.get("实际执行ID", 1)).value
        if not actual_id:
            continue
        actual_id = str(actual_id).strip()
        mapping[actual_id] = {
            "actual_id": actual_id,
            "actual_title": ws.cell(row, headers.get("实际执行标题", 2)).value or "",
            "auto_ids": split_ids(ws.cell(row, headers.get("匹配到Codex用例ID", 4)).value),
            "ft_ids": split_ids(ws.cell(row, headers.get("匹配到功能用例ID", 5)).value),
            "match_method": ws.cell(row, headers.get("匹配口径", 6)).value or "",
            "match_note": ws.cell(row, headers.get("匹配说明", 7)).value or "",
        }
    return mapping


def case_metadata(case_map: dict) -> dict[str, dict]:
    return {item["id"]: item for item in case_map.get("cases", [])}


def selected_ids(case_map: dict, profile: str) -> list[str]:
    profiles = case_map.get("profiles", {})
    selected = profiles.get(profile) or profiles.get(case_map.get("default_profile")) or profiles.get("mandatory") or {}
    return list(selected.get("core", [])) + list(selected.get("module", []))


def read_run_summary(args: argparse.Namespace, run_dir: Path | None) -> dict:
    if args.mark_core_only:
        return {}
    summary_path = Path(args.summary) if args.summary else (run_dir / "automation-run-summary.json" if run_dir else None)
    if not summary_path or not summary_path.exists():
        raise FileNotFoundError(f"Cannot find automation summary: {summary_path}")
    return load_json(summary_path)


def result_indexes(results: list[dict], execution_mapping: dict[str, dict]):
    by_actual = {item.get("id"): item for item in results if item.get("id")}
    by_auto = defaultdict(list)
    by_ft = defaultdict(list)
    for result in results:
        actual_id = result.get("id")
        mapped = execution_mapping.get(actual_id, {})
        for auto_id in mapped.get("auto_ids", []):
            by_auto[auto_id].append(result)
        for ft_id in mapped.get("ft_ids", []):
            by_ft[ft_id].append(result)
    return by_actual, by_auto, by_ft


def fill_main_sheet(ws, id_header: str, ids_to_mark: set[str], results_by_id: dict[str, list[dict]], run_time: str) -> None:
    header_row = find_header(ws, [id_header], 3)
    headers = ensure_columns(ws, header_row, [
        "每轮必跑",
        "每轮必跑原因",
        "本轮执行标记",
        "本轮执行结果",
        "本轮实际执行ID",
        "本轮证据目录",
        "本轮报告路径",
        "本轮执行时间",
    ])
    id_col = headers.get(id_header)
    if not id_col:
        return
    for row in range(header_row + 1, (ws.max_row or header_row) + 1):
        case_id = ws.cell(row, id_col).value
        if not case_id:
            continue
        case_id = str(case_id).strip()
        is_core = case_id in ids_to_mark
        row_results = results_by_id.get(case_id, [])
        ws.cell(row, headers["每轮必跑"]).value = "是" if is_core else ""
        if is_core and not ws.cell(row, headers["每轮必跑原因"]).value:
            ws.cell(row, headers["每轮必跑原因"]).value = "核心回归必跑：覆盖第一版上线核心路径或高频用户路径。"
        ws.cell(row, headers["本轮执行标记"]).value = "已执行" if row_results else ("待执行" if is_core else "")
        ws.cell(row, headers["本轮执行结果"]).value = aggregate_status(row_results)
        ws.cell(row, headers["本轮实际执行ID"]).value = "\n".join(sorted({item.get("id", "") for item in row_results if item.get("id")}))
        ws.cell(row, headers["本轮证据目录"]).value = "\n".join(sorted({item.get("case_dir", "") for item in row_results if item.get("case_dir")}))
        ws.cell(row, headers["本轮报告路径"]).value = "\n".join(sorted({item.get("case_report") or item.get("runner_report") or "" for item in row_results if item.get("case_report") or item.get("runner_report")}))
        ws.cell(row, headers["本轮执行时间"]).value = run_time if row_results else ""
    style_existing_used_area(ws, header_row)


def create_core_sheet(wb, case_map: dict, execution_mapping: dict[str, dict], mandatory_ids: list[str], summary: dict) -> None:
    ensure_sheet_removed(wb, "每轮必跑核心集")
    ws = wb.create_sheet("每轮必跑核心集", 0)
    headers = [
        "执行ID",
        "套件",
        "Runner",
        "优先级",
        "测试场景",
        "是否每轮必跑",
        "匹配Codex用例ID",
        "匹配功能用例ID",
        "必跑原因",
        "本轮是否执行",
        "本轮结果",
        "证据目录",
    ]
    ws.append(headers)
    mandatory_set = set(mandatory_ids)
    executed = {item.get("id"): item for item in summary.get("results", [])}
    for item in case_map.get("cases", []):
        actual_id = item.get("id")
        mapped = execution_mapping.get(actual_id, {})
        result = executed.get(actual_id, {})
        ws.append([
            actual_id,
            item.get("suite", ""),
            item.get("runner", ""),
            item.get("priority", ""),
            item.get("title", ""),
            "是" if actual_id in mandatory_set else "",
            "\n".join(mapped.get("auto_ids", [])) or "未匹配",
            "\n".join(mapped.get("ft_ids", [])) or "未匹配",
            item.get("must_run_reason", ""),
            "已执行" if result else "",
            normalize_status(result.get("status", "")) if result else "",
            result.get("case_dir", "") if result else "",
        ])
    style_sheet(ws)


def create_result_sheets(wb, summary: dict, execution_mapping: dict[str, dict]) -> None:
    ensure_sheet_removed(wb, "本轮执行结果")
    ensure_sheet_removed(wb, "本轮证据索引")

    result_ws = wb.create_sheet("本轮执行结果", 1)
    result_ws.append([
        "实际执行ID",
        "套件",
        "优先级",
        "测试场景",
        "执行结果",
        "操作步骤",
        "预期结果",
        "实际结果/失败原因",
        "测试结论",
        "匹配Codex用例ID",
        "匹配功能用例ID",
        "用例报告",
        "证据目录",
        "关键截图",
    ])
    for result in summary.get("results", []):
        mapped = execution_mapping.get(result.get("id"), {})
        steps = result.get("steps") or []
        step_text = "\n".join(
            f"{idx + 1}. {step.get('action', '')} - {normalize_status(step.get('status', '')) or step.get('status', '')}: {step.get('text', '')}"
            for idx, step in enumerate(steps)
        )
        screenshots = result.get("screenshots_flat") or []
        result_ws.append([
            result.get("id", ""),
            result.get("suite", ""),
            result.get("priority", ""),
            result.get("title", ""),
            normalize_status(result.get("status", "")),
            step_text,
            result.get("expected_result", "") or "见用例场景预期。",
            result.get("actual_result", "") or result.get("reason", "") or result.get("problem_description", ""),
            result.get("conclusion", "") or ("通过" if result.get("status") in {"passed", "pass"} else result.get("reason", "")),
            "\n".join(mapped.get("auto_ids", [])),
            "\n".join(mapped.get("ft_ids", [])),
            result.get("case_report") or result.get("runner_report") or "",
            result.get("case_dir", ""),
            "\n".join(screenshots[:10]),
        ])
    style_sheet(result_ws)

    evidence_ws = wb.create_sheet("本轮证据索引", 2)
    evidence_ws.append(["实际执行ID", "证据类型", "证据路径", "说明"])
    for result in summary.get("results", []):
        if result.get("case_report"):
            evidence_ws.append([result.get("id"), "case-report", result.get("case_report"), "单用例 Markdown 报告"])
        if result.get("runner_report"):
            evidence_ws.append([result.get("id"), "runner-report", result.get("runner_report"), "Runner 汇总 JSON"])
        if result.get("transcript_file"):
            evidence_ws.append([result.get("id"), "transcript", result.get("transcript_file"), "完整对话文本"])
        if result.get("reply_delta_file"):
            evidence_ws.append([result.get("id"), "reply-delta", result.get("reply_delta_file"), "本轮回复增量"])
        for screenshot in result.get("screenshots_flat") or []:
            evidence_ws.append([result.get("id"), "screenshot", screenshot, "截图证据"])
    style_sheet(evidence_ws)


def main() -> None:
    args = parse_args()
    root = Path(args.root).resolve()
    casebook = Path(args.casebook).resolve() if args.casebook else root / "PRD" / "Qbot_TestCase.xlsx"
    if not casebook.exists():
        raise FileNotFoundError(f"Casebook not found: {casebook}")
    run_dir = Path(args.run_dir).resolve() if args.run_dir else None
    output = Path(args.output).resolve() if args.output else (run_dir / "Qbot_TestCase_Result.xlsx" if run_dir else casebook)

    case_map = load_json(CASE_MAP_FILE)
    mandatory_ids = selected_ids(case_map, args.profile)
    wb = load_workbook(casebook)
    execution_mapping = load_execution_mapping(wb)
    mandatory_auto_ids = set()
    mandatory_ft_ids = set()
    for actual_id in mandatory_ids:
        mapped = execution_mapping.get(actual_id, {})
        mandatory_auto_ids.update(mapped.get("auto_ids", []))
        mandatory_ft_ids.update(mapped.get("ft_ids", []))

    summary = read_run_summary(args, run_dir)
    if summary:
        results = summary.get("results", [])
        _, by_auto, by_ft = result_indexes(results, execution_mapping)
        run_time = summary.get("ended_at", "")
    else:
        by_auto = {}
        by_ft = {}
        run_time = ""

    if "Codex自动化用例" in wb.sheetnames:
        fill_main_sheet(wb["Codex自动化用例"], "Codex用例ID", mandatory_auto_ids, by_auto, run_time)
    if "功能测试用例" in wb.sheetnames:
        fill_main_sheet(wb["功能测试用例"], "功能用例ID", mandatory_ft_ids, by_ft, run_time)

    create_core_sheet(wb, case_map, execution_mapping, mandatory_ids, summary)
    if summary:
        create_result_sheets(wb, summary, execution_mapping)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    if summary and run_dir:
        latest = root / "PRD" / "Qbot_TestCase_LastRun.xlsx"
        try:
            shutil.copy2(output, latest)
        except Exception:
            pass
    print(json.dumps({
        "status": "ok",
        "casebook": str(casebook),
        "output": str(output),
        "run_dir": str(run_dir) if run_dir else None,
        "mandatory_actual_cases": len(mandatory_ids),
        "mandatory_auto_cases_mapped": len(mandatory_auto_ids),
        "mandatory_ft_cases_mapped": len(mandatory_ft_ids),
        "result_count": len(summary.get("results", [])) if summary else 0,
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
