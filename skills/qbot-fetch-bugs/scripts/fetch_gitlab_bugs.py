#!/usr/bin/env python3
"""Fetch QBot/deepbankV2 bug issues from GitLab and export a styled Excel report."""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import urllib.error
import urllib.parse
import urllib.request
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ModuleNotFoundError:
    print(
        json.dumps(
            {
                "status": "blocked",
                "blocked_reason": "missing_python_dependency_openpyxl",
                "install": "python3 -m pip install --user openpyxl",
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    sys.exit(2)


DEFAULT_BASE_URL = "https://gitlab.daikuan.qihoo.net"
DEFAULT_PROJECT_PATH = "songrongxin/deepbankv2"
DEFAULT_TOKEN_FILE = Path.home() / ".config" / "qbot" / "gitlab-token"
TZ = ZoneInfo("Asia/Shanghai")

PRIORITY_ORDER = {"高": 1, "中": 2, "低": 3}
STATUS_ORDER = {"待解决": 1, "处理中": 2, "待验收": 3, "已解决": 4}

HIGH_PATTERNS = [
    r"crash|崩溃|crashloopbackoff|err_module_not_found",
    r"构建.*失败|部署.*失败|生产构建|无法启动|启动失败|打不开|未打开",
    r"登录失败|认证失败|鉴权失败|权限边界|数据丢失|安全",
    r"agent.*无响应|无有效回复|无法对话|chat stuck|stuck running|多轮上下文.*(不稳定|丢失|失败)",
    r"发送按钮.*(拦截|不可点击|无法发送)|输入区.*无法发送",
    r"模型无法感知|运行时注入断裂|路由到.*无法执行|raw ADK artifacts|冲掉进行中消息",
    r"control-plane|desktop-local|personal turn in docker",
    r"p0|p1",
]

MEDIUM_PATTERNS = [
    r"无法|不能|失败|不可用|不显示|缺少|无反应|点击无响应|无法确认|无法编辑|无法删除|无法修改|干扰用户",
    r"遮挡.*(消息|内容|按钮|入口|文案)|重叠.*(消息|内容|按钮|入口|文案)",
    r"安装失败|校验失败|回退版本后无反馈|重试.*丢失|duplicates|does not retry",
    r"附件|上传|设置|知识|连接器|项目|窗口|菜单栏|runtime family selector|model menu",
]

LOW_PATTERNS = [
    r"样式问题|布局问题|文案问题|间距|颜色|边框|视觉|未右对齐|文案被截断|内容布局错位",
]

MODULE_RULES = [
    ("登录鉴权与账号组织", "登录|认证|鉴权|oauth|lingxi|tokenless|token|账号|组织|用户身份|inherit another user's"),
    ("个人设置", "个人设置|用户画像|角色人设|界面主题|主题风格|默认执行方式|personal settings"),
    ("知识与成果库", "知识|knowledge|自然语言补充|知识源|知识引擎|本体|成果库|任务成果|artifact panel|artifacts|raw adk artifacts"),
    ("连接器与连应用", "连接器|connector|connectors|mcphub|mcp|连应用"),
    ("专家与角色", "专家|expert|experts|召唤|创建专家|我的专家|expert summon|selected expert"),
    ("技能与技能市场", "技能|skillhub|skills?|skill market|技能市场|安装失败|回退版本|更新技能|bsdtar|unzip|zip 解包"),
    ("会话输入区与附件", "composer|输入区|输入框|附件|上传|加号|发送按钮|foot-mask|model menu|安全级别|工作模式|craft|ask|plan"),
    ("任务与会话列表", "任务列表|新会话|任务搜索|会话列表|session switch|session .*title|会话切换|右键|编辑会话|删除会话|completion stats|sessions stay titled"),
    ("核心会话与 Agent 回复", "核心对话|会话页|agent-chat|assistant-ui|assistant chat|agent 回复|回复|消息气泡|多轮上下文|重新生成|regenerate|askuserquestion|chat stuck|无法使用 codex 会话|无法完成基础会话|用户消息|内部提示|diagnostics|chat text"),
    ("模型连接与运行时执行", "runtime|desktop-local|claude code|codex|llm|模型|anthropic|openai proxy|responses|native binary|control-plane|adk|resume handles|sdk|平台 llm|shell shim|codex\\.exe|claude|model/key|执行方式"),
    ("项目与工作区", "项目|project|工作区|workspace|项目运行时|project-runtime|项目资产|项目工作台|空间栏|空间功能"),
    ("桌面客户端安装与窗口", "electron|窗口|菜单栏|系统控件|titlebarstyle|dmg|release/mac|packaged|app\\.asar|spawn enotdir|数据根|%localappdata%|\\.deepbank|进程|single-instance|优雅停机|崩溃日志|主进程|preload|路径分隔符|路径白名单|路径脱敏|windows|linux|macos|macOS"),
    ("Teams 嵌入", "teams|teams360|360teams"),
    ("自动化测试与质量门禁", "e2e|自动化|playwright|测试|regression expects|uiux|单测|npm run check|质量门禁|静态|cross-sweep"),
    ("服务部署与可用性", "crashloopbackoff|生产构建|构建失败|部署失败|docker|k8s|health|服务启动|npm ci --omit|fflate|服务端启动"),
    ("通用界面布局", "样式|布局|遮挡|重叠|边框|文案|视觉|对齐|对比度|logo|侧栏|sidebar|esc|弹窗|概览入口|横向滚动|卡片边框"),
]

MODULE_SOURCE_BY_NAME = {
    "登录鉴权与账号组织": "src/App.tsx auth shell；server/auth.mjs",
    "个人设置": "src/AssistantConfig.tsx",
    "知识与成果库": "src/KnowledgeView.tsx；src/RoutingConfig.tsx；src/ArtifactPanel.tsx",
    "连接器与连应用": "src/ConnectorsView.tsx；src/ComposerTools.tsx 连应用",
    "专家与角色": "src/ExpertsView.tsx；server/expert-builder-*",
    "技能与技能市场": "src/ExpertsView.tsx 技能页；src/ComposerTools.tsx 技能选择；server/skill-*",
    "会话输入区与附件": "src/ComposerTools.tsx；src/ComposerExtras.tsx；src/components/assistant-ui/attachment.tsx",
    "任务与会话列表": "src/Sidebar.tsx SessionListItem；src/runtime.tsx session state",
    "核心会话与 Agent 回复": "src/components/assistant-ui/thread.tsx；src/Thread.tsx；server/engine.mjs",
    "模型连接与运行时执行": "src/runtime.tsx；electron/desktop-agent-host.mjs；server/engine.mjs",
    "项目与工作区": "src/ProjectsView.tsx；server/project-*",
    "桌面客户端安装与窗口": "electron/main.cjs；electron/preload.cjs；桌面包入口",
    "Teams 嵌入": "src/teams360/Root.tsx；electron/teams360-host-contract.md",
    "自动化测试与质量门禁": "test/e2e；scripts；QA 自动化链路",
    "服务部署与可用性": "server/*；server/Dockerfile；部署运行环境",
    "通用界面布局": "src/App.tsx；src/Sidebar.tsx；全局 CSS / 通用 UI 组件",
}

# Explicit overrides are based on previously reviewed issue content. These are product-module
# classifications, not GitLab area labels.
MODULE_OVERRIDES_BY_IID = {
    337: "服务部署与可用性",
    332: "核心会话与 Agent 回复",
    320: "会话输入区与附件",
    319: "模型连接与运行时执行",
    317: "自动化测试与质量门禁",
    314: "核心会话与 Agent 回复",
    268: "模型连接与运行时执行",
    256: "模型连接与运行时执行",
    303: "桌面客户端安装与窗口",
    326: "模型连接与运行时执行",
    309: "桌面客户端安装与窗口",
    308: "模型连接与运行时执行",
    307: "桌面客户端安装与窗口",
    306: "桌面客户端安装与窗口",
    305: "模型连接与运行时执行",
    304: "技能与技能市场",
    290: "模型连接与运行时执行",
    288: "模型连接与运行时执行",
    271: "模型连接与运行时执行",
    267: "核心会话与 Agent 回复",
    255: "模型连接与运行时执行",
    250: "任务与会话列表",
    231: "核心会话与 Agent 回复",
    225: "模型连接与运行时执行",
    223: "专家与角色",
    219: "模型连接与运行时执行",
    210: "技能与技能市场",
    204: "桌面客户端安装与窗口",
    202: "模型连接与运行时执行",
    198: "核心会话与 Agent 回复",
    194: "登录鉴权与账号组织",
    179: "模型连接与运行时执行",
    171: "模型连接与运行时执行",
    114: "知识与成果库",
    112: "模型连接与运行时执行",
    111: "核心会话与 Agent 回复",
    91: "核心会话与 Agent 回复",
    89: "桌面客户端安装与窗口",
    77: "自动化测试与质量门禁",
    72: "通用界面布局",
    343: "桌面客户端安装与窗口",
    333: "个人设置",
    331: "会话输入区与附件",
    330: "知识与成果库",
    339: "桌面客户端安装与窗口",
    341: "Teams 嵌入",
    321: "任务与会话列表",
    312: "个人设置",
    270: "会话输入区与附件",
    263: "技能与技能市场",
    262: "连接器与连应用",
    260: "任务与会话列表",
    259: "核心会话与 Agent 回复",
    253: "核心会话与 Agent 回复",
    248: "会话输入区与附件",
    247: "知识与成果库",
    239: "会话输入区与附件",
    238: "核心会话与 Agent 回复",
    236: "会话输入区与附件",
    235: "核心会话与 Agent 回复",
    234: "项目与工作区",
    233: "会话输入区与附件",
    232: "任务与会话列表",
    230: "任务与会话列表",
    229: "核心会话与 Agent 回复",
    221: "模型连接与运行时执行",
    220: "自动化测试与质量门禁",
    213: "任务与会话列表",
    209: "技能与技能市场",
    197: "专家与角色",
    195: "连接器与连应用",
    192: "技能与技能市场",
    122: "自动化测试与质量门禁",
    257: "核心会话与 Agent 回复",
}


def detect_qbot_test_agent_root() -> Path:
    env_root = os.environ.get("QBOT_TEST_AGENT_ROOT")
    if env_root:
        return Path(env_root).expanduser().resolve()

    cwd = Path.cwd().resolve()
    for candidate in [cwd, *cwd.parents]:
        if (candidate / "AGENTS.md").exists() and (candidate / "skills").exists():
            return candidate

    default = Path.home() / "Documents" / "QbotTestAgent"
    if default.exists():
        return default.resolve()

    return cwd


def read_token() -> tuple[str, str]:
    for env_name in ["GITLAB_TOKEN", "QBOT_FEEDBACK_GITLAB_TOKEN", "DEEPBANK_FEEDBACK_GITLAB_TOKEN"]:
        token = os.environ.get(env_name)
        if token:
            return token.strip(), f"env:{env_name}"

    for env_name in ["GITLAB_TOKEN_FILE", "QBOT_FEEDBACK_GITLAB_TOKEN_FILE", "DEEPBANK_FEEDBACK_GITLAB_TOKEN_FILE"]:
        token_file = os.environ.get(env_name)
        if token_file:
            path = Path(token_file).expanduser()
            if path.exists():
                return path.read_text(encoding="utf-8").strip(), str(path)

    if DEFAULT_TOKEN_FILE.exists():
        return DEFAULT_TOKEN_FILE.read_text(encoding="utf-8").strip(), str(DEFAULT_TOKEN_FILE)

    raise RuntimeError(
        "GitLab token not found. Configure GITLAB_TOKEN or write the token to ~/.config/qbot/gitlab-token."
    )


def api_get_json(url: str, token: str) -> tuple[list[dict[str, Any]], str]:
    request = urllib.request.Request(url, headers={"PRIVATE-TOKEN": token})
    try:
        with urllib.request.urlopen(request, timeout=45) as response:
            payload = json.loads(response.read().decode("utf-8"))
            next_page = response.headers.get("x-next-page", "")
            if not isinstance(payload, list):
                raise RuntimeError(f"Unexpected GitLab API payload: {type(payload).__name__}")
            return payload, next_page
    except urllib.error.HTTPError as exc:
        body = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"GitLab API {exc.code} {exc.reason}: {body[:500]}") from exc
    except urllib.error.URLError as exc:
        raise RuntimeError(f"GitLab API network error: {exc.reason}") from exc


def build_issues_url(base_url: str, project_path: str, page: int, extra: dict[str, str]) -> str:
    encoded_project = urllib.parse.quote(project_path, safe="")
    query = {
        "scope": "all",
        "state": "all",
        "per_page": "100",
        "page": str(page),
        "order_by": "updated_at",
        "sort": "desc",
        **extra,
    }
    return f"{base_url.rstrip('/')}/api/v4/projects/{encoded_project}/issues?{urllib.parse.urlencode(query)}"


def fetch_issue_set(base_url: str, project_path: str, token: str, extra: dict[str, str]) -> list[dict[str, Any]]:
    issues: list[dict[str, Any]] = []
    page = 1
    while True:
        url = build_issues_url(base_url, project_path, page, extra)
        data, next_page = api_get_json(url, token)
        issues.extend(data)
        if not next_page or not data:
            break
        page = int(next_page)
    return issues


def issue_title_has_bug(title: str) -> bool:
    return bool(re.search(r"(^|[^a-z0-9])bug([^a-z0-9]|$)|【\s*bug\s*】", title or "", flags=re.I))


def fetch_bug_issues(base_url: str, project_path: str, token: str) -> list[dict[str, Any]]:
    by_iid: dict[int, dict[str, Any]] = {}
    sources: defaultdict[int, set[str]] = defaultdict(set)

    for issue in fetch_issue_set(base_url, project_path, token, {"labels": "kind/bug"}):
        by_iid[int(issue["iid"])] = issue
        sources[int(issue["iid"])].add("kind/bug 标签")

    for issue in fetch_issue_set(base_url, project_path, token, {"search": "bug", "in": "title"}):
        iid = int(issue["iid"])
        if issue_title_has_bug(issue.get("title", "")):
            by_iid[iid] = issue
            sources[iid].add("标题包含 Bug")

    for iid, issue in by_iid.items():
        issue["_bug_match_source"] = "；".join(sorted(sources[iid])) or "未知"

    return list(by_iid.values())


def clean_text(value: str | None) -> str:
    text = value or ""
    text = re.sub(r"!\[[^\]]*\]\([^)]+\)", "", text)
    text = re.sub(r"\[([^\]]+)\]\([^)]+\)", r"\1", text)
    text = re.sub(r"[`*_>#-]", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def extract_short_summary(description: str | None) -> str:
    description = description or ""
    patterns = [
        r"\*\*Short summary\*\*:\s*([^\n]+)",
        r"### Problem\s+([\s\S]*?)(?:\n###|\n\*\*Observed|\n\*\*Expected|$)",
        r"问题描述[:：]\s*([^\n]+)",
        r"实际结果[:：]\s*([^\n]+)",
    ]
    for pattern in patterns:
        match = re.search(pattern, description, flags=re.I)
        if match:
            return clean_text(match.group(1))[:260]
    return clean_text(description)[:260]


def extract_actual_result(description: str | None) -> str:
    description = description or ""
    for pattern in [r"(?:\*\*)?Actual result(?:\*\*)?\s*:\s*([^\n]+)", r"实际结果[:：]\s*([^\n]+)"]:
        match = re.search(pattern, description, flags=re.I)
        if match:
            return clean_text(match.group(1))[:300]
    return ""


def classify_priority(issue: dict[str, Any]) -> tuple[str, str]:
    labels = issue.get("labels") or []
    lowered = {label.lower() for label in labels}
    if "priority/high" in lowered:
        return "高", "命中 GitLab 标签 priority/high"
    if "priority/medium" in lowered:
        return "中", "命中 GitLab 标签 priority/medium"
    if "priority/low" in lowered:
        return "低", "命中 GitLab 标签 priority/low"

    text = "\n".join(
        [
            issue.get("title", ""),
            extract_short_summary(issue.get("description")),
            extract_actual_result(issue.get("description")),
            clean_text(issue.get("description")),
        ]
    )
    if any(re.search(pattern, text, flags=re.I) for pattern in HIGH_PATTERNS):
        return "高", "无优先级标签；标题/正文命中核心阻断、生产崩溃、认证、Agent 回复或数据风险关键词"
    if any(re.search(pattern, text, flags=re.I) for pattern in MEDIUM_PATTERNS):
        return "中", "无优先级标签；标题/正文命中功能不可用、交互阻塞、重要入口异常或附件/设置类关键词"
    if any(re.search(pattern, text, flags=re.I) for pattern in LOW_PATTERNS):
        return "低", "无优先级标签；主要为样式、文案、视觉展示类低风险问题"
    return "低", "无优先级标签；未命中核心阻断或主要功能异常关键词，按低优先级待分诊"


def classify_module(issue: dict[str, Any]) -> tuple[str, str]:
    iid = int(issue.get("iid", 0))
    if iid in MODULE_OVERRIDES_BY_IID:
        module_name = MODULE_OVERRIDES_BY_IID[iid]
        source = MODULE_SOURCE_BY_NAME.get(module_name, "产品功能入口")
        return module_name, f"按 issue 标题/正文的实际用户场景显式归类，未使用 GitLab area 标签；源码参照：{source}"

    text = "\n".join(
        [
            issue.get("title", ""),
            extract_short_summary(issue.get("description")),
            extract_actual_result(issue.get("description")),
            clean_text(issue.get("description")),
        ]
    )
    for module_name, pattern in MODULE_RULES:
        if re.search(pattern, text, flags=re.I):
            source = MODULE_SOURCE_BY_NAME.get(module_name, "产品功能入口")
            return module_name, f"按实际功能入口分类；匹配问题标题/正文摘要/实际结果，源码参照：{source}"
    return "未明确归属", "未使用 GitLab 标签；标题/正文摘要/实际结果未命中明确产品功能入口，需要人工二次分诊"


def classify_status(issue: dict[str, Any]) -> tuple[str, str]:
    labels = [str(label) for label in (issue.get("labels") or [])]
    lowered = [label.lower() for label in labels]
    if issue.get("state") == "closed":
        return "已解决", "GitLab issue 已关闭"
    if any(re.search(r"status/(done|resolved|verified|closed|fixed)", label) for label in lowered):
        return "已解决", "GitLab 状态标签显示已完成/已验证/已修复"
    if any(re.search(r"status/(in-review|review|verification|verify|qa)", label) for label in lowered):
        return "待验收", "GitLab 状态标签显示待评审/待验证"
    if any(re.search(r"status/(in-progress|doing|wip)", label) for label in lowered):
        return "处理中", "GitLab 状态标签显示处理中"
    if any(re.search(r"status/(ready|todo|open)", label) for label in lowered):
        return "待解决", "GitLab 状态标签显示待处理"
    return "待解决", f"GitLab state={issue.get('state')}，未发现更具体的 status 标签"


def to_local_datetime(value: str | None) -> str:
    if not value:
        return ""
    try:
        normalized = value.replace("Z", "+00:00")
        dt = datetime.fromisoformat(normalized)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(TZ).strftime("%Y-%m-%d %H:%M:%S")
    except ValueError:
        return value


def analyze_issues(issues: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows = []
    for issue in issues:
        priority, priority_basis = classify_priority(issue)
        module_name, module_basis = classify_module(issue)
        status, status_basis = classify_status(issue)
        rows.append(
            {
                "iid": int(issue.get("iid", 0)),
                "title": issue.get("title", ""),
                "url": issue.get("web_url", ""),
                "bug_match_source": issue.get("_bug_match_source", ""),
                "priority": priority,
                "module_name": module_name,
                "status": status,
                "resolved": "是" if status == "已解决" else "否",
                "status_detail": status_basis,
                "labels": ", ".join(issue.get("labels") or []),
                "assignees": ", ".join(
                    user.get("username") or user.get("name") or "" for user in (issue.get("assignees") or [])
                )
                or "未指派",
                "author": (issue.get("author") or {}).get("username") or (issue.get("author") or {}).get("name") or "",
                "created_at": to_local_datetime(issue.get("created_at")),
                "updated_at": to_local_datetime(issue.get("updated_at")),
                "summary": extract_short_summary(issue.get("description")),
                "actual_result": extract_actual_result(issue.get("description")),
                "priority_basis": priority_basis,
                "module_basis": module_basis,
                "state": issue.get("state", ""),
            }
        )

    rows.sort(
        key=lambda row: (
            PRIORITY_ORDER.get(row["priority"], 99),
            STATUS_ORDER.get(row["status"], 99),
            -row["iid"],
        )
    )
    return rows


def set_sheet_title(ws, title: str, subtitle: str, end_col: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    ws.cell(1, 1, title)
    ws.cell(1, 1).fill = PatternFill("solid", fgColor="16324F")
    ws.cell(1, 1).font = Font(bold=True, color="FFFFFF", size=15)
    ws.cell(1, 1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
    ws.cell(2, 1, subtitle)
    ws.cell(2, 1).fill = PatternFill("solid", fgColor="EAF2F8")
    ws.cell(2, 1).font = Font(color="1F2937", size=10)
    ws.cell(2, 1).alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[2].height = 34


def style_range(ws, min_row: int, max_row: int, min_col: int, max_col: int, header_row: int | None = None) -> None:
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.font = Font(color="111827", size=10)

    if header_row:
        for cell in ws.iter_rows(min_row=header_row, max_row=header_row, min_col=min_col, max_col=max_col).__next__():
            cell.fill = PatternFill("solid", fgColor="D9EAD3")
            cell.font = Font(bold=True, color="0F172A", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def add_table(ws, name: str, ref: str) -> None:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def set_widths(ws, widths: list[float]) -> None:
    for index, width in enumerate(widths, start=1):
        column = get_column_letter(index)
        ws.column_dimensions[column].width = width


def add_overview_sheet(wb: Workbook, rows: list[dict[str, Any]], fetched_at: str, source_url: str) -> None:
    ws = wb.create_sheet("统计总览")
    set_sheet_title(ws, "DeepbankV2 Bug 当前情况总览", f"数据源：{source_url}；抓取时间：{fetched_at}；状态为 GitLab 实时状态。", 8)
    total = len(rows)
    priority_counter = Counter(row["priority"] for row in rows)
    status_counter = Counter(row["status"] for row in rows)
    unresolved = total - status_counter.get("已解决", 0)

    ws.append([])
    ws.append(["指标", "数量", "说明", "", "指标", "数量", "说明", ""])
    ws.append(["Bug issue 总数", total, "kind/bug 标签或标题包含 Bug 的 issue 去重后数量", "", "高优先级", priority_counter.get("高", 0), "核心阻断、生产/构建、认证、核心对话等", ""])
    ws.append(["未解决/处理中/待验收", unresolved, "除已解决状态外", "", "中优先级", priority_counter.get("中", 0), "重要功能、交互阻塞、可见异常", ""])
    ws.append(["已解决", status_counter.get("已解决", 0), "GitLab 已关闭或状态标签显示完成", "", "低优先级", priority_counter.get("低", 0), "样式、文案、轻量展示类", ""])
    style_range(ws, 4, 7, 1, 8, header_row=4)
    set_widths(ws, [24, 10, 48, 4, 12, 10, 48, 4])
    for row in range(4, 8):
        ws.row_dimensions[row].height = 30

    start = 10
    ws.cell(start, 1, "优先级")
    ws.cell(start, 2, "Issue数量")
    for offset, (key, count) in enumerate(sorted(priority_counter.items(), key=lambda item: PRIORITY_ORDER.get(item[0], 99)), start=1):
        ws.cell(start + offset, 1, key)
        ws.cell(start + offset, 2, count)
    style_range(ws, start, start + max(len(priority_counter), 1), 1, 2, header_row=start)

    ws.cell(start, 4, "状态")
    ws.cell(start, 5, "Issue数量")
    for offset, (key, count) in enumerate(sorted(status_counter.items(), key=lambda item: STATUS_ORDER.get(item[0], 99)), start=1):
        ws.cell(start + offset, 4, key)
        ws.cell(start + offset, 5, count)
    style_range(ws, start, start + max(len(status_counter), 1), 4, 5, header_row=start)

    module_counter = Counter(row["module_name"] for row in rows)
    ws.cell(start, 7, "模块")
    ws.cell(start, 8, "Issue数量")
    for offset, (key, count) in enumerate(module_counter.most_common(), start=1):
        ws.cell(start + offset, 7, key)
        ws.cell(start + offset, 8, count)
    style_range(ws, start, start + max(len(module_counter), 1), 7, 8, header_row=start)
    ws.freeze_panes = "A4"


def add_detail_sheet(wb: Workbook, rows: list[dict[str, Any]], fetched_at: str, source_url: str) -> None:
    ws = wb.create_sheet("Bug明细")
    headers = [
        "Issue ID",
        "Bug标题",
        "Bug链接",
        "Bug命中来源",
        "Bug优先级",
        "Bug所属模块",
        "Bug最新状态",
        "是否已解决",
        "状态依据",
        "GitLab标签",
        "指派人",
        "作者",
        "创建时间",
        "更新时间",
        "正文摘要",
        "实际结果摘要",
        "优先级依据",
        "模块依据",
        "GitLab原始state",
    ]
    set_sheet_title(ws, "DeepbankV2 Bug Issue 最新状态明细", f"数据源：{source_url}；抓取时间：{fetched_at}；链接列为可点击 GitLab issue。", len(headers))
    ws.append([])
    ws.append(headers)
    for row in rows:
        ws.append(
            [
                f"#{row['iid']}",
                row["title"],
                row["url"],
                row["bug_match_source"],
                row["priority"],
                row["module_name"],
                row["status"],
                row["resolved"],
                row["status_detail"],
                row["labels"],
                row["assignees"],
                row["author"],
                row["created_at"],
                row["updated_at"],
                row["summary"],
                row["actual_result"],
                row["priority_basis"],
                row["module_basis"],
                row["state"],
            ]
        )

    max_row = ws.max_row
    style_range(ws, 4, max_row, 1, len(headers), header_row=4)
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:S{max_row}"
    set_widths(ws, [10, 50, 56, 18, 12, 24, 13, 12, 34, 42, 18, 15, 20, 20, 54, 46, 48, 48, 15])

    for row_index in range(5, max_row + 1):
        link_cell = ws.cell(row_index, 3)
        if link_cell.value:
            link_cell.hyperlink = link_cell.value
            link_cell.style = "Hyperlink"

        priority = ws.cell(row_index, 5).value
        status = ws.cell(row_index, 7).value
        if priority == "高":
            ws.cell(row_index, 5).fill = PatternFill("solid", fgColor="F4CCCC")
        elif priority == "中":
            ws.cell(row_index, 5).fill = PatternFill("solid", fgColor="FFF2CC")
        elif priority == "低":
            ws.cell(row_index, 5).fill = PatternFill("solid", fgColor="D9EAD3")
        if status == "已解决":
            ws.cell(row_index, 7).fill = PatternFill("solid", fgColor="D9EAD3")
        elif status == "处理中":
            ws.cell(row_index, 7).fill = PatternFill("solid", fgColor="DDEBF7")
        elif status == "待验收":
            ws.cell(row_index, 7).fill = PatternFill("solid", fgColor="EADCF8")
        else:
            ws.cell(row_index, 7).fill = PatternFill("solid", fgColor="FCE4D6")

        ws.row_dimensions[row_index].height = 48

    if max_row >= 5:
        add_table(ws, "BugDetailsTable", f"A4:S{max_row}")


def add_rules_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("分类规则")
    set_sheet_title(ws, "Bug 分类规则说明", "用于审计本次统计口径；模块归属不使用 GitLab area 标签。", 4)
    rows = [
        ["分类项", "规则", "说明", ""],
        ["Bug范围", "kind/bug 标签或标题显式包含 bug / Bug / 【Bug】", "两类来源取并集并按 issue IID 去重。", ""],
        ["状态", "closed/done/resolved/verified=已解决；in-progress=处理中；review/verification=待验收；ready/open=待解决", "以 GitLab 当前 state 和 status 标签为准。", ""],
        ["优先级：高", "priority/high 或核心阻断关键词", "包含生产崩溃、构建/部署失败、认证失败、核心会话无响应、数据/权限风险等。", ""],
        ["优先级：中", "priority/medium 或重要功能/交互异常关键词", "包含无法编辑/删除、不显示、遮挡重要入口、附件状态不可确认等。", ""],
        ["优先级：低", "priority/low 或纯样式/文案/轻微展示问题", "不影响主流程且可绕过的问题归低优先级。", ""],
        ["模块归属", "按用户实际遇到问题的功能入口分类", "参考 QBot 源码入口，如 Sidebar、ComposerTools、ExpertsView、KnowledgeView、ConnectorsView、ProjectsView、AssistantConfig、Electron、Runtime、Server 等。", ""],
    ]
    for row in rows:
        ws.append(row)
    style_range(ws, 4, ws.max_row, 1, 4, header_row=4)
    set_widths(ws, [20, 58, 82, 4])
    for row_index in range(4, ws.max_row + 1):
        ws.row_dimensions[row_index].height = 42
    ws.freeze_panes = "A5"


def add_source_sheet(
    wb: Workbook,
    base_url: str,
    project_path: str,
    source_url: str,
    fetched_at: str,
    output_path: Path,
    token_source: str,
    rows: list[dict[str, Any]],
) -> None:
    ws = wb.create_sheet("数据源信息")
    set_sheet_title(ws, "数据源信息", "用于审计本次统计的数据来源和抓取参数。", 4)
    values = [
        ["字段", "值", "备注", ""],
        ["GitLab Base URL", base_url, "接口来源", ""],
        ["项目路径", project_path, "GitLab 项目路径", ""],
        ["项目地址", source_url, "Issue 来源", ""],
        ["API筛选", "labels=kind/bug；search=bug,in=title；state=all,scope=all", "两类结果去重合并", ""],
        ["抓取时间", fetched_at, "Asia/Shanghai", ""],
        ["Bug issue 数", len(rows), "见 Bug明细 sheet", ""],
        ["输出文件", str(output_path), "本地 QA 产物", ""],
        ["Token 来源", token_source, "不写入 token 明文", ""],
        ["说明", "本表不修改 deepbankV2 仓库，仅用于 QA 统计。", "", ""],
    ]
    for row in values:
        ws.append(row)
    style_range(ws, 4, ws.max_row, 1, 4, header_row=4)
    set_widths(ws, [22, 88, 42, 4])
    ws.freeze_panes = "A5"


def build_workbook(
    rows: list[dict[str, Any]],
    output_path: Path,
    fetched_at: str,
    base_url: str,
    project_path: str,
    token_source: str,
) -> None:
    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)
    source_url = f"{base_url.rstrip('/')}/{project_path}"

    add_overview_sheet(wb, rows, fetched_at, source_url)
    add_detail_sheet(wb, rows, fetched_at, source_url)
    add_rules_sheet(wb)
    add_source_sheet(wb, base_url, project_path, source_url, fetched_at, output_path, token_source, rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Fetch deepbankV2 bug issues and export timestamped Excel.")
    parser.add_argument("--base-url", default=os.environ.get("GITLAB_BASE_URL", DEFAULT_BASE_URL))
    parser.add_argument("--project-path", default=os.environ.get("DEEPBANK_PROJECT_PATH", DEFAULT_PROJECT_PATH))
    parser.add_argument("--output-dir", default="")
    parser.add_argument("--write-json", action="store_true", help="Also write raw/classified JSON files next to the Excel.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    root = detect_qbot_test_agent_root()
    output_dir = Path(args.output_dir).expanduser().resolve() if args.output_dir else root / "bug"
    timestamp = datetime.now(TZ).strftime("%Y%m%d%H%M")
    fetched_at = datetime.now(TZ).strftime("%Y-%m-%d %H:%M:%S")
    output_path = output_dir / f"Qbot_Bug_Issues_{timestamp}.xlsx"
    if output_path.exists():
        suffix = 2
        while True:
            candidate = output_dir / f"Qbot_Bug_Issues_{timestamp}-{suffix}.xlsx"
            if not candidate.exists():
                output_path = candidate
                break
            suffix += 1

    try:
        token, token_source = read_token()
        issues = fetch_bug_issues(args.base_url, args.project_path, token)
        rows = analyze_issues(issues)
        build_workbook(rows, output_path, fetched_at, args.base_url, args.project_path, token_source)

        if args.write_json:
            output_dir.mkdir(parents=True, exist_ok=True)
            (output_dir / f"Qbot_Bug_Issues_{timestamp}.raw.json").write_text(
                json.dumps(issues, ensure_ascii=False, indent=2), encoding="utf-8"
            )
            (output_dir / f"Qbot_Bug_Issues_{timestamp}.classified.json").write_text(
                json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8"
            )

        summary = {
            "status": "ok",
            "xlsxPath": str(output_path),
            "issueCount": len(rows),
            "priority": dict(Counter(row["priority"] for row in rows)),
            "latestStatus": dict(Counter(row["status"] for row in rows)),
            "outputDir": str(output_dir),
            "fetchedAt": fetched_at,
        }
        print(json.dumps(summary, ensure_ascii=False, indent=2))
        return 0
    except Exception as exc:
        print(
            json.dumps(
                {
                    "status": "failed",
                    "error": str(exc),
                    "outputDir": str(output_dir),
                    "hint": "Check GitLab token/network access, then rerun.",
                },
                ensure_ascii=False,
                indent=2,
            )
        )
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
