#!/usr/bin/env python3
"""Apply framework-readable case corrections and append #668/#669 SIT cases."""

from copy import copy
from pathlib import Path
import sys

from openpyxl import load_workbook


DEFAULT = Path(__file__).resolve().parents[1] / "PRD" / "QBot系统SIT自动化测试用例_框架清零版_2026-07-11.xlsx"
SOURCE_REV = "deepbankV2 origin/main@43767488658c9ad0e521349ac09f5d6af323a966"


def headers(ws):
    return {str(ws.cell(1, col).value).strip(): col for col in range(1, ws.max_column + 1) if ws.cell(1, col).value}


def row_for_id(ws, case_id):
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, 1).value == case_id:
            return row
    raise KeyError(case_id)


def update_case(ws, case_id, values):
    cols = headers(ws)
    row = row_for_id(ws, case_id)
    for name, value in values.items():
        ws.cell(row, cols[name]).value = value


def append_case(ws, values):
    cols = headers(ws)
    row = ws.max_row + 1
    template = row - 1
    for col in range(1, ws.max_column + 1):
        src, dst = ws.cell(template, col), ws.cell(row, col)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        dst.alignment = copy(src.alignment)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)
        dst.font = copy(src.font)
        dst.protection = copy(src.protection)
    ws.row_dimensions[row].height = ws.row_dimensions[template].height
    for name, value in values.items():
        ws.cell(row, cols[name]).value = value


COMMON_PRECONDITION = (
    "QBot 使用包含 #668/#669 的最新 main 本地代码启动并已登录工作台；"
    "自动化仅通过用户可见 UI 与 runner 控制面观察代理执行，不修改 deepbankV2 源码。"
)
COMMON_EVIDENCE = (
    "保留技能操作前卡片、操作后 feedback、已安装/市场终态截图；"
    "框架同时保存真实点击、请求命中次数和页面文本。数据 fixture 缺失时必须可信阻塞，不得伪造通过。"
)
SELECTORS = (
    '[data-testid="nav-experts"]; [data-testid="skills-tab"]; text=技能市场; '
    '.skill-search input; .skill-card; .skill-install; [data-testid="skill-operation-feedback"]'
)


NEW_CASES = [
    {
        "用例ID": "SIT-SKILL-027", "优先级": "P0", "产品模块": "技能功能", "子功能": "拒装显式重试",
        "测试场景": "#668 可重试拒装技能应允许用户显式重试且一次点击只触发一次安装请求",
        "前置条件": COMMON_PRECONDITION + " SkillHub/本机 overlay 准备 installRetryable=true 的拒装技能 qa-runtime-retryable。",
        "测试数据": "自动化技能标识=qa-runtime-retryable",
        "执行入口/Selector": SELECTORS + '; text=重试安装',
        "执行步骤": "1. 搜索 qa-runtime-retryable。\n2. 确认卡片显示【重试安装】且可点击。\n3. runner 只观察 POST /api/skills/install，不改写响应。\n4. 单击一次【重试安装】。\n5. 等待成功或明确失败终态并统计请求次数。",
        "预期结果": "显式重试入口可用；一次点击只产生一次安装请求；结果收敛且提示可理解。",
        "成功判定": "按钮可点、request_hits=1，feedback 离开安装中并显示成功/失败/未就绪原因。",
        "失败判定": "按钮置灰/缺失、一次点击请求为 0 或大于 1、一直安装中或无结果均失败。",
        "证据要求": COMMON_EVIDENCE, "自动化Runner": "Playwright UI Agent / control-plane observe", "执行层级": "SIT", "每轮必跑": "是",
        "来源ID": "#668", "来源类型": SOURCE_REV, "备注": "新增：覆盖用户显式单次重试通道，不把失败终态误判为框架失败。",
    },
    {
        "用例ID": "SIT-SKILL-028", "优先级": "P0", "产品模块": "技能功能", "子功能": "审计拒装终态",
        "测试场景": "#668 同版本 audit_terminal 技能执行通用自动对账时不得隐式发起安装",
        "前置条件": COMMON_PRECONDITION + " 准备同版本 audit_terminal 拒装技能 qa-audit-terminal。",
        "测试数据": "自动化技能标识=qa-audit-terminal",
        "执行入口/Selector": SELECTORS + '; [data-testid="nav-settings-menu"]; [data-testid="nav-settings"]; [data-testid="assistant-reconcile-skills"]',
        "执行步骤": "1. 搜索并记录 qa-audit-terminal 的拒装状态。\n2. 进入个人设置点击【立即对账技能】（不传 userRetrySlugs）。\n3. runner 观察 /api/skills/install 请求。\n4. 返回市场复查同一技能。",
        "预期结果": "通用自动对账不触发安装请求；同版本审计拒装状态保持且有说明。",
        "成功判定": "request_hits=0；卡片仍显示装不上/拒装/重试安装等明确状态。",
        "失败判定": "自动对账触发安装请求、状态静默丢失或重复安装均失败。",
        "证据要求": COMMON_EVIDENCE, "自动化Runner": "Playwright UI Agent / control-plane observe", "执行层级": "SIT", "每轮必跑": "是",
        "来源ID": "#668", "来源类型": SOURCE_REV, "备注": "新增：保护 #486 的自动对账零安装请求契约。",
    },
    {
        "用例ID": "SIT-SKILL-029", "优先级": "P0", "产品模块": "技能功能", "子功能": "拒装卸载闭环",
        "测试场景": "#668 卸载拒装/未就绪技能后应清除本机 overlay 并恢复市场可安装状态",
        "前置条件": COMMON_PRECONDITION + " 已安装列表准备带拒装/未就绪本机状态的 qa-uninstall-rejected。",
        "测试数据": "自动化技能标识=qa-uninstall-rejected",
        "执行入口/Selector": SELECTORS + '; text=已安装; .skill-del; [data-testid="skill-uninstall-confirm"]',
        "执行步骤": "1. 在已安装列表定位 qa-uninstall-rejected 并确认拒装/未就绪标识。\n2. 点击删除并确认。\n3. 等待删除弹窗关闭。\n4. 到技能市场搜索同一标识。",
        "预期结果": "卸载完成；市场卡恢复可安装；旧装不上/拒装 overlay 不残留。",
        "成功判定": "市场存在同名卡、安装按钮可用，且卡片不再显示装不上/拒装。",
        "失败判定": "PG 已删但本机 overlay 仍标红、技能仍在已安装列表或市场不可安装均失败。",
        "证据要求": COMMON_EVIDENCE, "自动化Runner": "Playwright UI Agent / ui", "执行层级": "SIT", "每轮必跑": "是",
        "来源ID": "#668", "来源类型": SOURCE_REV, "备注": "新增：验证卸载同时清安装记录、缓存和本机投影状态。",
    },
    {
        "用例ID": "SIT-SKILL-030", "优先级": "P0", "产品模块": "技能功能", "子功能": "必填依赖级联",
        "测试场景": "#669 安装主技能时应递归先安装全部未安装必填依赖并在反馈中列出",
        "前置条件": COMMON_PRECONDITION + " SkillHub 准备根技能及两个可成功安装的 required 依赖。",
        "测试数据": "自动化技能标识=qa-dep-root-success\ndependencies=qa-dep-leaf-a,qa-dep-leaf-b",
        "执行入口/Selector": SELECTORS,
        "执行步骤": "1. 搜索 qa-dep-root-success。\n2. 点击安装并等待终态。\n3. 读取 feedback 的“并级联安装”列表。\n4. 进入已安装，逐一核对根技能和两个依赖。",
        "预期结果": "依赖先于主技能安装；反馈列出级联依赖；三张技能卡均出现在已安装。",
        "成功判定": "feedback 成功且含“并级联安装”；根技能和 dependencies 中全部标识均可见。",
        "失败判定": "只装主技能、漏装依赖、反馈不列依赖、依赖失败仍写根技能均失败。",
        "证据要求": COMMON_EVIDENCE, "自动化Runner": "Playwright UI Agent / ui", "执行层级": "SIT", "每轮必跑": "是",
        "来源ID": "#669", "来源类型": SOURCE_REV, "备注": "新增：递归 required 依赖成功主链路。",
    },
    {
        "用例ID": "SIT-SKILL-031", "优先级": "P1", "产品模块": "技能功能", "子功能": "已安装依赖跳过",
        "测试场景": "#669 必填依赖已安装时应跳过且不重复写入/重复展示为本次级联",
        "前置条件": COMMON_PRECONDITION + " qa-dep-leaf-existing 已安装；根技能 qa-dep-root-existing 依赖它且尚未安装。",
        "测试数据": "自动化技能标识=qa-dep-root-existing\ndependencies=qa-dep-leaf-existing",
        "执行入口/Selector": SELECTORS + '; text=已安装',
        "执行步骤": "1. 在已安装确认 qa-dep-leaf-existing。\n2. 安装 qa-dep-root-existing。\n3. 读取 feedback。\n4. 复查根技能和依赖卡片。",
        "预期结果": "主技能成功；已有依赖跳过，不在本次“并级联安装”反馈中重复出现。",
        "成功判定": "前置依赖存在；主技能成功；feedback 不把已有依赖列为新级联；两者终态均存在。",
        "失败判定": "重复安装/重复记录已有依赖，或因依赖已安装而阻断主技能均失败。",
        "证据要求": COMMON_EVIDENCE, "自动化Runner": "Playwright UI Agent / ui", "执行层级": "SIT", "每轮必跑": "否",
        "来源ID": "#669", "来源类型": SOURCE_REV, "备注": "新增：dependencyMatchesInstall 跳过分支。",
    },
    {
        "用例ID": "SIT-SKILL-032", "优先级": "P0", "产品模块": "技能功能", "子功能": "依赖失败阻断",
        "测试场景": "#669 任一必填依赖安装失败时应点名失败依赖且主技能不得写入已安装",
        "前置条件": COMMON_PRECONDITION + " qa-dep-root-failure 依赖一个确定安装失败/审计拒装的 qa-dep-leaf-failure。",
        "测试数据": "自动化技能标识=qa-dep-root-failure\ndependencies=qa-dep-leaf-failure",
        "执行入口/Selector": SELECTORS,
        "执行步骤": "1. 搜索并安装 qa-dep-root-failure。\n2. 等待失败 feedback。\n3. 核对文案点名依赖技能。\n4. 进入已安装搜索主技能。",
        "预期结果": "显示依赖技能安装失败及失败依赖；主技能不出现在已安装。",
        "成功判定": "feedback 为失败且含“依赖技能/失败”和依赖标识；rootInstalled=false。",
        "失败判定": "依赖失败后主技能仍写入、错误未点名依赖或无可理解原因均失败。",
        "证据要求": COMMON_EVIDENCE, "自动化Runner": "Playwright UI Agent / ui", "执行层级": "SIT", "每轮必跑": "是",
        "来源ID": "#669", "来源类型": SOURCE_REV, "备注": "新增：failedDependency 与主技能写入阻断。",
    },
    {
        "用例ID": "SIT-SKILL-033", "优先级": "P0", "产品模块": "技能功能", "子功能": "循环依赖保护",
        "测试场景": "#669 必填依赖形成循环时应 fail-closed 并阻止主技能安装",
        "前置条件": COMMON_PRECONDITION + " qa-dep-root-cycle 与其依赖技能形成 required 循环引用。",
        "测试数据": "自动化技能标识=qa-dep-root-cycle\ndependencies=qa-dep-cycle-b",
        "执行入口/Selector": SELECTORS,
        "执行步骤": "1. 搜索 qa-dep-root-cycle。\n2. 点击安装并等待 feedback。\n3. 核对循环引用提示。\n4. 进入已安装确认主技能不存在。",
        "预期结果": "明确提示必填依赖循环引用；主技能不写入已安装。",
        "成功判定": "feedback 失败且含循环引用/循环依赖；rootInstalled=false。",
        "失败判定": "死循环/超时、静默失败、主技能仍安装或暴露内部堆栈均失败。",
        "证据要求": COMMON_EVIDENCE, "自动化Runner": "Playwright UI Agent / ui", "执行层级": "SIT", "每轮必跑": "是",
        "来源ID": "#669", "来源类型": SOURCE_REV, "备注": "新增：required_skill_dependency_cycle 产品化终态。",
    },
]


UPDATES = {
    "SIT-HOME-008": {"备注": "2026-07-14：使用 connector-only 专用 runner；发送前断言手动连接器仍选中，禁止通用 reset 清空选择。"},
    "SIT-HOME-023": {"备注": "2026-07-14：停止后同时采集 bridge.running、cancelPending 和停止按钮消失状态，最长等待 60 秒。"},
    "SIT-HOME-025": {"执行入口/Selector": '[data-testid="composer-input"]; [data-testid="composer-send"]; runner-control-plane-proxy:/api/desktop-agent/turn-context', "备注": "2026-07-14：改为 runner 自有控制面代理注入 503；不依赖 CDP preload Node 上下文。"},
    "SIT-HOME-030": {"执行入口/Selector": '[data-testid="composer-feedback"]; [data-testid="quick-feedback-panel"]; [data-testid="quick-feedback-submit"]; runner-control-plane-proxy:/api/feedback-issues/intake', "备注": "2026-07-14：代理以 dry-run 捕获真实脱敏 payload，避免创建外部 issue。"},
    "SIT-HOME-037": {"测试数据": "testflies/qbot-image-test.png（1200×720，包含报名100/到场70/成交12和漏斗图）", "备注": "2026-07-14：替换 1×1 像素占位图，要求识别指标和漏斗含义。"},
    "SIT-HOME-038": {"测试数据": "testflies/qbot-image-test.png, qbot-image-flow.png, qbot-image-risk.png（三张不同内容的1200×720 PNG）", "备注": "2026-07-14：真正上传三张内容互异图片，验证跨图综合。"},
    "SIT-HOME-039": {"测试数据": "testflies/qbot-image-test.png + testflies/qbot-data.json", "备注": "2026-07-14：明确混合上传真实业务图片和 JSON，不再只用占位图。"},
    "SIT-HOME-056": {"执行入口/Selector": '.aui-attachment-root; .aui-attachment-tile-remove', "备注": "2026-07-14：先 hover 目标附件卡，再点击真实隐藏删除按钮。"},
    "SIT-EXPERT-012": {"执行入口/Selector": '.exp-recent-item; .exp-recent-del', "备注": "2026-07-14：先 hover 最近召唤卡片，再识别/点击隐藏移除按钮。"},
    "SIT-SKILL-013": {"执行入口/Selector": '.skill-card; [data-testid="nav-settings-menu"]; [data-testid="nav-settings"]; [data-testid="assistant-reconcile-skills"]; [data-testid="assistant-reconcile-result"]', "备注": "2026-07-14：卡片无操作入口时走个人设置“立即对账技能”，不再只等待。"},
    "SIT-SKILL-021": {"执行入口/Selector": '.skill-install; runner-control-plane-proxy:/api/skills/install', "备注": "2026-07-14：runner 控制面代理返回可控安装失败并验证无脏已安装状态。"},
    "SIT-CONN-012": {"执行入口/Selector": '[data-testid="composer-connectors-menu"]; runner-control-plane-proxy:/api/capabilities', "备注": "2026-07-14：代理变换 capabilities 中已选连接器为 needs_auth，并在重启后复查警告。"},
    "SIT-CONN-013": {"执行入口/Selector": '[data-testid="connectors-refresh"]; runner-control-plane-proxy:/api/connectors/catalog?refresh=force', "备注": "2026-07-14：代理注入刷新失败，断言已有缓存卡片保留。"},
}


def main():
    target = Path(sys.argv[1]).expanduser().resolve() if len(sys.argv) > 1 else DEFAULT
    wb = load_workbook(target)
    by_id = {}
    for ws in wb.worksheets:
        if ws.cell(1, 1).value == "用例ID":
            for row in range(2, ws.max_row + 1):
                by_id[ws.cell(row, 1).value] = ws
    for case_id, values in UPDATES.items():
        update_case(by_id[case_id], case_id, values)

    skill = wb["技能功能"]
    existing = {skill.cell(row, 1).value for row in range(2, skill.max_row + 1)}
    for case in NEW_CASES:
        if case["用例ID"] not in existing:
            append_case(skill, case)

    summary = wb["汇总说明"]
    summary["B2"] = "2026-07-14"
    summary["B3"] = SOURCE_REV
    summary["E6"], summary["F6"], summary["G6"], summary["H6"] = 33, 14, 14, 5
    summary["I6"], summary["J6"], summary["K6"] = 14, 9, 7
    summary["B10"] = "有条件通过：共149条；新增 #668/#669 的7条技能功能用例。自动化 fixture 缺失必须可信阻塞，框架不得以未执行动作判通过。"

    changes = wb["审查变更清单"]
    for case in NEW_CASES:
        row = changes.max_row + 1
        template = row - 1
        for col in range(1, changes.max_column + 1):
            changes.cell(row, col)._style = copy(changes.cell(template, col)._style)
            changes.cell(row, col).alignment = copy(changes.cell(template, col).alignment)
        values = ["新增", case["用例ID"], "技能功能", case["测试场景"], case["来源ID"], "是"]
        for col, value in enumerate(values, 1):
            changes.cell(row, col).value = value

    if wb.calculation is not None:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    wb.save(target)
    print(target)
    functional = sum(wb[name].max_row - 1 for name in ("首页会话组合", "专家功能", "技能功能", "连接器功能", "成果项目"))
    print(f"技能用例={skill.max_row - 1}; 功能全量={functional}")


if __name__ == "__main__":
    main()
