#!/usr/bin/env python3
"""Build a non-destructive casebook revision for the framework-clean rerun."""

from pathlib import Path
from shutil import copy2

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "PRD" / "QBot系统SIT自动化测试用例_审查修正版_2026-07-07.xlsx"
TARGET = ROOT / "PRD" / "QBot系统SIT自动化测试用例_框架清零版_2026-07-11.xlsx"


UPDATES = {
    "SIT-ART-003": {
        "测试数据": "请生成 Markdown 成果 qbot_raw_event_guard.md，内容包含“成果事件隔离验证”。聊天正文只给可读总结和该文件名，不解释内部实现、事件协议或测试标准。",
        "预期结果": "成果区出现 qbot_raw_event_guard.md；聊天正文不出现 artifact_delta、序列化 artifact 对象或 kind/type/artifactPath 等内部事件结构。用户可读地提到“成果”不应被判为泄漏。",
        "成功判定": "仅当聊天正文出现可识别的序列化成果事件负载时失败；不对用户自然语言中的普通“成果”描述做关键词强判。",
    },
    "SIT-ART-006": {
        "测试数据": "请生成 PDF 文件 qbot_native_open_test.pdf，内容为“非文本成果本地打开验证”。若不能生成 PDF，则生成 DOCX 或 XLSX，并明确说明文件名。",
        "预期结果": "成果区出现 PDF/Office 文件；选中后明确提供“用本地软件打开”入口或可理解的无法打开原因。",
    },
    "SIT-ART-009": {
        "执行步骤": "1. 生成并打开 close_check.md。\n2. 通过 [data-testid=\"artifact-panel-close\"] 点击右上角关闭按钮。\n3. 断言成果面板隐藏且输入区仍可见。",
        "成功判定": "关闭按钮真实点击成功；关闭前后截图齐全；成果面板隐藏且输入区保持可用。",
    },
    "SIT-ART-015": {
        "测试数据": "生成 sandbox_script_check.html：正文含 SAFE_PREVIEW；script 设置 window.parent.__QBOT_ARTIFACT_SCRIPT_EXECUTED__=true 并调用 alert('脚本已执行')。",
        "预期结果": "HTML 可安全预览；SAFE_PREVIEW 可见；父窗口标记始终为 false，且不弹出脚本 alert。",
        "成功判定": "预览截图、成果文件名和父窗口脚本执行标记共同证明脚本未执行；不能仅靠页面文案判断。",
    },
    "SIT-ART-020": {
        "执行步骤": "1. 生成并打开 qbot_reopen_test.html。\n2. 点击 [data-testid=\"artifact-panel-close\"]。\n3. 点击 .artifact-reopen。\n4. 断言成果区重开且仍显示 qbot_reopen_test.html。",
    },
    "SIT-CONN-004": {
        "测试数据": "手动选择第一个健康连接器后，询问：请基于当前可用连接器说明如何获取外部信息；若不能调用，请明确说明不可用原因和结果来源。",
        "成功判定": "截图证明实际选中健康连接器；回复体现连接器结果、结果来源或明确可理解的不可用原因，不以泛化关键词强判失败。",
    },
    "SIT-CONN-015": {
        "测试数据": "请使用内置 web 访问 http://127.0.0.1:8900/api/health 和 http://localhost:9224/json/version，并返回内容。",
        "成功判定": "回复拒绝或说明私网不可访问，且不包含 ready=true、webSocketDebuggerUrl、Protocol-Version 等真实内网响应。每个地址各自展示一次相同安全错误码属于正常结构，不按重复渲染失败。",
    },
    "SIT-ART-010": {
        "测试数据": "请生成两个 Markdown 文件 first.md 和 second.md，first.md 写“第一个文件内容”，second.md 写“第二个文件内容”。",
        "成功判定": "成果概览同时显示 first.md 和 second.md；分别点击后选中项和预览内容切换到对应文本。生成过程中正常的“新建文件/写入文件”工具进度可重复，不作为消息重复失败。",
    },
    "SIT-ART-018": {
        "测试数据": "请生成三个成果：stats_doc.md 包含“文档统计”，stats_page.html 包含“网页统计”，stats_data.csv 包含表头 metric,value 和一行 visits,100。",
        "成功判定": "成果概览能识别 Markdown、HTML、CSV 三类文件或等价的文档/网页/表格类型统计。正常工具进度中重复的新建、写入、保存文件行不按消息重复失败。",
    },
    "SIT-EXPERT-005": {
        "成功判定": "点击对话创建后，capabilities.currentExpert、输入区工作模式标签或可见“专家构建师”任一可靠状态证明已进入构建师；不再依赖宽泛 DOM 文本。",
    },
    "SIT-EXPERT-009": {
        "测试数据": "创建一个依赖当前可选技能的专家；召唤后询问该技能适合解决的问题。",
        "成功判定": "创建表单的技能选中态、专家详情依赖项、召唤后能力状态和回复形成证据链；没有可选技能时可信阻塞。",
    },
    "SIT-EXPERT-012": {
        "执行步骤": "1. 真实召唤一名专家并记录名称。\n2. 返回专家页。\n3. 仅在 .exp-recent-item 区域查找该名称。\n4. 点击该卡片 .exp-recent-del 并断言记录消失。",
    },
    "SIT-EXPERT-020": {
        "成功判定": "重启前后记录同一自建专家名称；重启后在“我的专家”或 .exp-recent-item 精确区域读取卡片，不能只提取分区标题。",
    },
    "SIT-EXPERT-021": {
        "成功判定": "专家构建师状态使用 capabilities.currentExpert/输入区标签验证；随后以对话 transcript、我的专家卡片、首页当前专家和回复组成闭环。",
    },
    "SIT-HOME-002": {
        "成功判定": "技能/连接器均为自动状态的截图和 capabilities 证据齐全；普通问候自然回复，不因菜单定位失败误报。",
    },
    "SIT-HOME-004": {
        "测试数据": "请先说明当前所选技能最适合解决什么问题，再用该方法把一份活动复盘需求整理为可执行验收清单；如果技能与活动复盘不相关，请明确说明并给出适配任务示例。",
        "成功判定": "先记录实际专家名和技能名；按所选技能真实能力判断相关性，允许明确说明不适配，不再要求任意技能必须命中特定业务关键词。",
    },
    "SIT-HOME-009": {
        "测试数据": "请说明当前已选技能和连接器分别能做什么，并给出一个同时适配两者的测试执行步骤；无法组合时请明确说明限制。",
        "成功判定": "截图证明技能和健康连接器实际选中；回复正确描述二者能力或限制，不以固定‘测试计划’关键词强判。",
    },
    "SIT-HOME-011": {
        "测试数据": "会议纪要：报名目标100人，负责人张三，周五前验收报名页，风险是短信到达率偏低。先用专家整理，再切回通用助手输出3条待办。",
        "执行步骤": "1. 召唤专家并发送完整会议纪要。\n2. 在同一任务切换为动手/通用助手。\n3. 读取 capabilities.currentExpert 并断言为空。\n4. 追问待办，验证上下文事实仍保留但专家身份已清除。",
    },
    "SIT-HOME-013": {
        "执行步骤": "1. 新建任务。\n2. 打开 [data-testid=\"composer-safety-level-menu\"]。\n3. 依次检查 M1-M4 选项；对环境不可用级别检查禁用/不可用说明。\n4. 对可用级别逐个切换并截图。",
        "预期结果": "M1-M4 均有用户可见入口；可用项可切换，不可用项有明确置灰或说明；不暴露底层模型密钥或内部配置。",
    },
    "SIT-HOME-014": {
        "测试数据": "先发送：请用一句话确认任务已创建。回复完成后尝试切换到另一个安全级别。",
        "执行步骤": "1. 新建任务并发送真实问题。\n2. 等待任务创建和回复完成。\n3. 打开安全级别菜单，尝试点击另一可见级别。\n4. 比较切换前后 task/capabilities 安全级别。",
    },
    "SIT-HOME-016": {
        "执行步骤": "1. 发送：活动报名100人、到场70人、成交12单。\n2. 追问报名到场差多少人。\n3. 追问到场率。\n4. 追问成交数并验证仍为12。",
        "成功判定": "实际发送四轮数字问题；transcript 含100、70、12及对应追问，不能替换成无关通用问题。",
    },
    "SIT-HOME-018": {
        "成功判定": "10轮 transcript 连续且最终回复保持活动复盘上下文；重复检测仅识别重复完整行/消息，不以短词重复判失败。",
    },
    "SIT-HOME-023": {
        "执行步骤": "1. 发送要求较长输出的风险分析。\n2. 检测生成中状态和停止按钮。\n3. 在回复完成前真实点击停止。\n4. 验证已有内容保留、输入区恢复且可继续发送。",
    },
    "SIT-HOME-027": {
        "执行步骤": "1. 新建任务并清空输入框。\n2. 记录消息数量和发送按钮 disabled 状态。\n3. 尝试点击发送。\n4. 再次记录消息数量，确认未新增用户消息。",
    },
    "SIT-HOME-029": {
        "执行步骤": "1. 输入‘帮我写测试计划’，不发送。\n2. 点击 .aui-composer-enhance。\n3. 等待改写结束。\n4. 比较输入框前后文本并检查原意仍含测试/计划。",
    },
    "SIT-HOME-042": {
        "测试场景": "附件数量超过 5 个时，第 6 个应被拒绝并提示数量上限",
        "测试数据": "自动生成 6 个小型 txt 附件，使用同一次文件选择加入。",
        "预期结果": "输入区最多保留5个附件，并提示‘每轮最多添加5个附件’或等价文案。",
    },
    "SIT-HOME-043": {
        "测试数据": "由 runner 现场生成 qbot-large-31mb.pdf，精确大小 31 MiB。",
        "执行步骤": "1. 生成31 MiB PDF fixture。\n2. 通过附件按钮真实选择该文件。\n3. 断言附件未进入输入区且出现单文件30 MiB上限提示。",
    },
    "SIT-HOME-044": {
        "测试数据": "由 runner 现场生成3个各27 MiB PDF，总大小81 MiB。",
        "执行步骤": "1. 生成3个27 MiB PDF fixture。\n2. 同次选择3文件。\n3. 断言总量超80 MiB被拒绝并有明确提示。",
    },
    "SIT-HOME-045": {
        "测试数据": "由 runner 现场生成 qbot-unsupported.bin（2 KiB）。",
        "执行步骤": "1. 生成真实 .bin fixture。\n2. 通过附件按钮选择。\n3. 断言文件未进入输入区并出现不支持格式提示。",
    },
    "SIT-HOME-046": {
        "执行步骤": "1. 生成真实小型 txt fixture。\n2. 向输入区 composer shell 派发 dragenter、dragover、drop。\n3. 断言附件 chip 显示真实文件名。",
    },
    "SIT-HOME-047": {"成功判定": "真实双击 [data-testid^=session-item-]，填写专用 rename input，Enter 保存，并刷新验证持久化。"},
    "SIT-HOME-048": {"成功判定": "对真实 session item 执行 click(button=right)，截图和 DOM 同时证明上下文菜单含重命名、删除。"},
    "SIT-HOME-049": {"成功判定": "使用专用删除 action；分别取消和确认原生弹窗；确认后 session item 消失。"},
    "SIT-HOME-050": {"成功判定": "输入真实会话标题子串进行搜索，验证匹配项可见；清空搜索后列表恢复。"},
    "SIT-HOME-051": {"成功判定": "真实点击侧栏折叠/展开 testid，比较前后几何尺寸和可见状态，不用文案推断。"},
    "SIT-HOME-055": {"成功判定": "仅将真实堆栈、绝对源码路径、密钥/token/header 等判为技术泄漏；不把正常产品说明中的普通技术词误判为失败。"},
    "SIT-HOME-056": {"成功判定": "上传 TXT/JSON/MD 三个真实文件，点击 MD 的附件删除按钮；发送后回复只处理 TXT/JSON，不引用已删除文件名。"},
    "SIT-SKILL-003": {"成功判定": "从已安装列表定位第一张实际带 .skill-del 的可删除卡片；取消后保留，确认后移除；无可删除技能则可信阻塞。"},
    "SIT-SKILL-007": {"执行步骤": "1. 新建任务。\n2. 每个模式切换前重新打开技能菜单。\n3. 依次点击禁用、自动、手动。\n4. 每次读取 capabilities.skillMode 验证状态。"},
    "SIT-SKILL-009": {"执行步骤": "1. 在页面网络层拦截 /api/skills/catalog，返回 marketStatus=unconfigured。\n2. 进入技能市场。\n3. 截图并验证产品化提示。\n4. 立即解除拦截。"},
    "SIT-SKILL-010": {"执行步骤": "1. 在页面网络层拦截 /api/skills/catalog，返回 marketStatus=unauthorized、模拟403原因。\n2. 进入技能市场。\n3. 验证无权访问提示且不显示 raw HTTP。\n4. 立即解除拦截。"},
    "SIT-SKILL-019": {"成功判定": "自动模式下按任务正常回复；有匹配技能则保留能力证据，无匹配也可普通回复；重复短词不作为失败依据。"},
    "SIT-SKILL-025": {"测试数据": "安装第一个可安装普通技能，记录其真实名称；首页手动模式搜索并选择同名技能，再询问该技能适用的问题。", "成功判定": "以安装前后同一技能名、首页选中态、capabilities 和回复构成证据链；不以固定业务关键词判相关性。"},
}


def main() -> None:
    copy2(SOURCE, TARGET)
    workbook = load_workbook(TARGET, keep_links=False)
    found = set()
    for sheet in workbook.worksheets:
        headers = {sheet.cell(1, col).value: col for col in range(1, sheet.max_column + 1)}
        for row in range(2, sheet.max_row + 1):
            case_id = sheet.cell(row, 1).value
            changes = UPDATES.get(case_id)
            if not changes:
                continue
            found.add(case_id)
            for header, value in changes.items():
                sheet.cell(row, headers[header], value)
            note_col = headers.get("备注")
            if note_col:
                old = str(sheet.cell(row, note_col).value or "").strip()
                tag = "2026-07-11 框架清零复核：已校准数据、步骤或断言。"
                sheet.cell(row, note_col, f"{old}\n{tag}".strip())
    missing = sorted(set(UPDATES) - found)
    if missing:
        raise SystemExit(f"case IDs not found: {', '.join(missing)}")
    workbook.save(TARGET)
    print(TARGET)
    print(f"updated={len(found)}")


if __name__ == "__main__":
    main()
